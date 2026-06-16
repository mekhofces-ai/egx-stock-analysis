from __future__ import annotations

import argparse
import hashlib
import json
import logging
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import pandas as pd
from sqlalchemy import and_, func, select
from sqlalchemy.orm import Session

from app.config import REPORT_TIMEZONE, RISK_NOTE, get_settings
from app.database import SessionLocal, init_db, sqlite_write_lock
from app.models import (
    JobsLog,
    NotificationLog,
    RecommendationEvaluation,
    RecommendationItem,
    RecommendationReport,
    Stock,
)
from app.services.daily_loss_audit import (
    EVAL_DATA_MISSING,
    EVAL_EVALUATED,
    EVAL_ENTRY_NOT_REACHED,
    EVAL_EXPIRED,
    EVAL_NOT_EVALUATED,
    EVAL_STOP_HIT,
    EVAL_TARGET_HIT,
    LOW_MISSING_DATA,
    classify_recommendation_path,
    _candles_after_recommendation,
    _latest_signal_price_before,
)


logger = logging.getLogger(__name__)
CAIRO_TZ = ZoneInfo(REPORT_TIMEZONE)
PROJECT_ROOT = Path(__file__).resolve().parents[2]
REPORT_DIR = PROJECT_ROOT / "reports" / "recommendation_performance"
REPORT_DIR.mkdir(parents=True, exist_ok=True)

FINAL_STATUSES = {EVAL_TARGET_HIT, EVAL_STOP_HIT, EVAL_EXPIRED}
NON_ACCURACY_STATUSES = {EVAL_NOT_EVALUATED, EVAL_DATA_MISSING, EVAL_ENTRY_NOT_REACHED}
HORIZONS = (1, 3, 5, 10)


def cairo_now() -> datetime:
    return datetime.now(CAIRO_TZ)


def _day_bounds(day: date) -> tuple[datetime, datetime]:
    start = datetime(day.year, day.month, day.day)
    return start, start + timedelta(days=1)


def _safe_float(value: Any) -> float | None:
    if value is None:
        return None
    try:
        number = float(value)
    except Exception:
        return None
    if pd.isna(number):
        return None
    return number


def _is_accuracy_row(status: Any) -> bool:
    return str(status or "").upper() not in NON_ACCURACY_STATUSES


def _details_source(details: dict[str, Any] | None, fallback: str | None = None) -> str | None:
    details = details if isinstance(details, dict) else {}
    source = details.get("source") or details.get("strategy_source") or fallback
    if not source and isinstance(details.get("strategy"), dict):
        source = details["strategy"].get("label")
    return str(source) if source else fallback


def _telegram_source(details: dict[str, Any] | None) -> str | None:
    details = details if isinstance(details, dict) else {}
    telegram = details.get("telegram")
    if isinstance(telegram, dict):
        value = telegram.get("top_channels") or telegram.get("label") or telegram.get("reason")
    else:
        value = details.get("telegram_source")
    if value is None:
        return None
    if isinstance(value, (list, tuple, set)):
        value = ", ".join(str(item) for item in value if str(item).strip())
    elif isinstance(value, dict):
        value = json.dumps(value, ensure_ascii=False, default=str)
    return str(value)[:255] if str(value).strip() else None


def _market_regime(details: dict[str, Any] | None) -> str | None:
    details = details if isinstance(details, dict) else {}
    market = details.get("market_evaluation") or details.get("market")
    if isinstance(market, dict):
        return market.get("market_regime") or market.get("regime")
    return details.get("market_regime")


def _horizon_frame(candles: pd.DataFrame, days: int) -> pd.DataFrame:
    if candles is None or candles.empty or "datetime" not in candles.columns:
        return pd.DataFrame()
    frame = candles.copy()
    frame["datetime"] = pd.to_datetime(frame["datetime"], errors="coerce")
    frame = frame.dropna(subset=["datetime"]).sort_values("datetime")
    dates = list(dict.fromkeys(frame["datetime"].dt.date.tolist()))
    selected_dates = set(dates[:days])
    return frame[frame["datetime"].dt.date.isin(selected_dates)].copy()


def _evaluate_path(
    db: Session,
    item: RecommendationItem,
    report: RecommendationReport,
    *,
    expiry_days: int = 10,
) -> tuple[Any, dict[str, Any], pd.DataFrame, str, str]:
    target_date = report.report_time.date()
    candles, evaluation_quality, timeframe_used = _candles_after_recommendation(db, item.symbol, report.report_time, target_date)
    signal_price = _latest_signal_price_before(db, item.symbol, report.report_time)
    final_audit = classify_recommendation_path(
        candles,
        entry_zone_low=item.entry_zone_low,
        entry_zone_high=item.entry_zone_high,
        stop_loss=item.stop_loss,
        target_1=item.target_1,
        target_2=item.target_2,
        target_3=item.target_3,
        signal=item.signal,
        evaluation_quality=evaluation_quality,
        signal_price=signal_price,
    )
    horizons: dict[str, Any] = {}
    for horizon in HORIZONS:
        frame = _horizon_frame(candles, horizon)
        if frame.empty:
            horizons[f"{horizon}d"] = {
                "status": EVAL_DATA_MISSING if evaluation_quality == LOW_MISSING_DATA else EVAL_NOT_EVALUATED,
                "days_evaluated": 0,
                "notes": "No future candle for this horizon yet.",
            }
            continue
        horizon_audit = classify_recommendation_path(
            frame,
            entry_zone_low=item.entry_zone_low,
            entry_zone_high=item.entry_zone_high,
            stop_loss=item.stop_loss,
            target_1=item.target_1,
            target_2=item.target_2,
            target_3=item.target_3,
            signal=item.signal,
            evaluation_quality=evaluation_quality,
            signal_price=signal_price,
        )
        horizons[f"{horizon}d"] = {
            "status": horizon_audit.evaluation_status,
            "quality": horizon_audit.final_quality,
            "days_evaluated": horizon_audit.days_evaluated,
            "latest_close": horizon_audit.close_after_recommendation,
            "actual_return_pct": horizon_audit.actual_return,
            "max_favorable_move_pct": horizon_audit.max_favorable_move_pct,
            "max_adverse_move_pct": horizon_audit.max_adverse_move_pct,
            "target_hit": bool(horizon_audit.target_1_hit or horizon_audit.target_2_hit or horizon_audit.target_3_hit),
            "stop_hit": bool(horizon_audit.stop_loss_hit),
            "notes": horizon_audit.root_cause,
        }
    if final_audit.evaluation_status == EVAL_EVALUATED and final_audit.days_evaluated >= expiry_days:
        final_audit.evaluation_status = EVAL_EXPIRED
        final_audit.final_quality = "Weak"
        final_audit.root_cause = f"{final_audit.root_cause} Evaluation expired after {expiry_days} trading day(s) without target/stop resolution."
    return final_audit, horizons, candles, evaluation_quality, timeframe_used


def _upsert_evaluation(
    db: Session,
    item: RecommendationItem,
    report: RecommendationReport,
    audit: Any,
    horizons: dict[str, Any],
    *,
    evaluation_quality: str,
    timeframe_used: str,
    evaluated_at: datetime,
) -> RecommendationEvaluation:
    row = db.scalar(select(RecommendationEvaluation).where(RecommendationEvaluation.recommendation_item_id == item.id))
    details = item.details_json if isinstance(item.details_json, dict) else {}
    target_hit = bool(audit.target_1_hit or audit.target_2_hit or audit.target_3_hit)
    stop_hit = bool(audit.stop_loss_hit)
    data = {
        "report_id": report.id,
        "symbol": item.symbol,
        "report_type": report.report_type,
        "recommendation_datetime": report.report_time,
        "recommendation_stage": item.signal,
        "strategy_source": _details_source(details, report.report_type),
        "telegram_source": _telegram_source(details),
        "market_regime": _market_regime(details),
        "evaluated_at": evaluated_at.replace(tzinfo=None),
        "days_evaluated": int(audit.days_evaluated or 0),
        "signal_price": audit.signal_price,
        "next_available_open": audit.next_available_open,
        "latest_close": audit.close_after_recommendation,
        "highest_after_signal": audit.max_price_after_signal,
        "lowest_after_signal": audit.min_price_after_signal,
        "actual_return_pct": audit.actual_return,
        "max_favorable_move_pct": audit.max_favorable_move_pct,
        "max_adverse_move_pct": audit.max_adverse_move_pct,
        "target_hit": target_hit,
        "stop_hit": stop_hit,
        "final_status": audit.evaluation_status,
        "final_quality": audit.final_quality,
        "evaluation_quality": evaluation_quality,
        "evaluation_notes": audit.root_cause,
        "horizons_json": horizons,
        "details_json": {
            "timeframe_used": timeframe_used,
            "entry_touched": audit.entry_touched,
            "actual_entry_price": audit.actual_entry_price,
            "time_to_target_minutes": audit.time_to_target_minutes,
            "time_to_stop_minutes": audit.time_to_stop_minutes,
            "result": audit.result,
            "risk_note": RISK_NOTE,
        },
        "updated_at": evaluated_at.replace(tzinfo=None),
    }
    if row is None:
        row = RecommendationEvaluation(recommendation_item_id=item.id, **data)
        db.add(row)
    else:
        for key, value in data.items():
            setattr(row, key, value)
    return row


def _candidate_rows(
    db: Session,
    *,
    as_of_date: date,
    include_today: bool = False,
    limit: int = 2000,
    reevaluate_final: bool = False,
) -> list[tuple[RecommendationItem, RecommendationReport, Stock | None, RecommendationEvaluation | None]]:
    day_start, _day_end = _day_bounds(as_of_date)
    cutoff = day_start if not include_today else datetime.now()
    cutoff_condition = RecommendationReport.report_time < cutoff if not include_today else RecommendationReport.report_time <= cutoff
    stmt = (
        select(RecommendationItem, RecommendationReport, Stock, RecommendationEvaluation)
        .join(RecommendationReport, RecommendationItem.report_id == RecommendationReport.id)
        .outerjoin(Stock, Stock.symbol == RecommendationItem.symbol)
        .outerjoin(RecommendationEvaluation, RecommendationEvaluation.recommendation_item_id == RecommendationItem.id)
        .where(cutoff_condition)
        .order_by(RecommendationReport.report_time.asc(), RecommendationItem.id.asc())
        .limit(limit)
    )
    rows = list(db.execute(stmt).all())
    filtered: list[tuple[RecommendationItem, RecommendationReport, Stock | None, RecommendationEvaluation | None]] = []
    for item, report, stock, evaluation in rows:
        if not reevaluate_final and evaluation and evaluation.final_status in FINAL_STATUSES:
            continue
        if evaluation and evaluation.final_status in {EVAL_NOT_EVALUATED, EVAL_EVALUATED, EVAL_ENTRY_NOT_REACHED}:
            filtered.append((item, report, stock, evaluation))
        elif evaluation is None:
            filtered.append((item, report, stock, evaluation))
    return filtered


def run_daily_re_evaluation(
    *,
    as_of_date: date | None = None,
    include_today: bool = False,
    limit: int = 2000,
    persist: bool = True,
    db: Session | None = None,
) -> dict[str, Any]:
    as_of_date = as_of_date or cairo_now().date()
    evaluated_at = cairo_now()

    def _run(active_db: Session) -> dict[str, Any]:
        job = JobsLog(job_name="recommendation_re_evaluation", status="running", started_at=evaluated_at.replace(tzinfo=None))
        if persist:
            active_db.add(job)
            active_db.flush()
        rows = _candidate_rows(active_db, as_of_date=as_of_date, include_today=include_today, limit=limit)
        updated: list[dict[str, Any]] = []
        skipped: list[dict[str, Any]] = []
        for item, report, stock, _evaluation in rows:
            try:
                audit, horizons, _candles, quality, timeframe = _evaluate_path(active_db, item, report)
                if audit.evaluation_status in {EVAL_NOT_EVALUATED, EVAL_DATA_MISSING}:
                    skipped.append({"symbol": item.symbol, "recommendation_item_id": item.id, "reason": audit.root_cause, "status": audit.evaluation_status})
                if persist:
                    row = _upsert_evaluation(
                        active_db,
                        item,
                        report,
                        audit,
                        horizons,
                        evaluation_quality=quality,
                        timeframe_used=timeframe,
                        evaluated_at=evaluated_at,
                    )
                    active_db.flush()
                    evaluation_id = row.id
                else:
                    evaluation_id = None
                updated.append(
                    {
                        "evaluation_id": evaluation_id,
                        "recommendation_item_id": item.id,
                        "report_id": report.id,
                        "report_type": report.report_type,
                        "recommendation_date": report.report_time,
                        "symbol": item.symbol,
                        "stock_name": stock.name if stock else item.company_name,
                        "recommendation_stage": item.signal,
                        "entry_from": item.entry_zone_low,
                        "entry_to": item.entry_zone_high,
                        "stop_loss": item.stop_loss,
                        "target_1": item.target_1,
                        "target_2": item.target_2,
                        "signal_price": audit.signal_price,
                        "latest_close": audit.close_after_recommendation,
                        "highest_after_signal": audit.max_price_after_signal,
                        "lowest_after_signal": audit.min_price_after_signal,
                        "actual_return_pct": audit.actual_return,
                        "max_favorable_move_pct": audit.max_favorable_move_pct,
                        "max_adverse_move_pct": audit.max_adverse_move_pct,
                        "days_evaluated": audit.days_evaluated,
                        "final_status": audit.evaluation_status,
                        "final_quality": audit.final_quality,
                        "evaluation_notes": audit.root_cause,
                    }
                )
            except Exception as exc:
                logger.exception("Recommendation re-evaluation failed for %s item %s", item.symbol, item.id)
                skipped.append({"symbol": item.symbol, "recommendation_item_id": item.id, "reason": str(exc), "status": "ERROR"})
        summary = summarize_evaluations(updated)
        if persist:
            job.status = "success"
            job.finished_at = cairo_now().replace(tzinfo=None)
            job.details = json.dumps({"summary": summary, "skipped": skipped[:100]}, ensure_ascii=False, default=str)
            active_db.commit()
        return {"as_of_date": as_of_date.isoformat(), "updated": updated, "skipped": skipped, "summary": summary}

    if db is not None:
        return _run(db)
    init_db(seed=True)
    with sqlite_write_lock():
        with SessionLocal() as active_db:
            return _run(active_db)


def summarize_evaluations(rows: list[dict[str, Any]]) -> dict[str, Any]:
    total = len(rows)
    evaluated = [row for row in rows if _is_accuracy_row(row.get("final_status"))]
    not_evaluated = sum(1 for row in rows if row.get("final_status") == EVAL_NOT_EVALUATED)
    missing = sum(1 for row in rows if row.get("final_status") == EVAL_DATA_MISSING)
    entry_not_reached = sum(1 for row in rows if row.get("final_status") == EVAL_ENTRY_NOT_REACHED)
    target_hits = [row for row in evaluated if row.get("final_status") == EVAL_TARGET_HIT or row.get("target_hit")]
    stop_hits = [row for row in evaluated if row.get("final_status") == EVAL_STOP_HIT or row.get("stop_hit")]
    open_rows = [row for row in rows if row.get("final_status") == EVAL_EVALUATED]
    returns = [_safe_float(row.get("actual_return_pct")) for row in evaluated]
    returns = [value for value in returns if value is not None]
    best = max(evaluated, key=lambda row: _safe_float(row.get("actual_return_pct")) if _safe_float(row.get("actual_return_pct")) is not None else -999999, default=None)
    worst = min(evaluated, key=lambda row: _safe_float(row.get("actual_return_pct")) if _safe_float(row.get("actual_return_pct")) is not None else 999999, default=None)
    reliable = len(evaluated) >= 5
    win_rate = round((len(target_hits) / len(evaluated)) * 100.0, 2) if reliable else None
    return {
        "total_recommendations": total,
        "evaluated_recommendations": len(evaluated),
        "not_evaluated_recommendations": not_evaluated,
        "missing_data_recommendations": missing,
        "entry_not_reached_recommendations": entry_not_reached,
        "target_hits_today": len(target_hits),
        "stop_hits_today": len(stop_hits),
        "open_recommendations": len(open_rows),
        "win_rate_pct": win_rate,
        "accuracy_note": "" if reliable else "Accuracy is not reliable yet because evaluated sample size is below 5.",
        "average_return_pct": round(sum(returns) / len(returns), 2) if returns else None,
        "best_stock_today": f"{best.get('symbol')} ({best.get('actual_return_pct')}%)" if best else "-",
        "worst_stock_today": f"{worst.get('symbol')} ({worst.get('actual_return_pct')}%)" if worst else "-",
    }


def _all_evaluations_frame(db: Session, *, limit: int = 5000) -> pd.DataFrame:
    rows = db.execute(
        select(RecommendationEvaluation, RecommendationItem, RecommendationReport, Stock)
        .join(RecommendationItem, RecommendationEvaluation.recommendation_item_id == RecommendationItem.id)
        .join(RecommendationReport, RecommendationEvaluation.report_id == RecommendationReport.id)
        .outerjoin(Stock, Stock.symbol == RecommendationEvaluation.symbol)
        .order_by(RecommendationEvaluation.updated_at.desc(), RecommendationEvaluation.id.desc())
        .limit(limit)
    ).all()
    data: list[dict[str, Any]] = []
    for ev, item, report, stock in rows:
        data.append(
            {
                "Report Date": report.report_time.date().isoformat(),
                "Recommendation Date": report.report_time.isoformat(sep=" ", timespec="seconds"),
                "Stock Symbol": ev.symbol,
                "Stock Name": stock.name if stock else item.company_name,
                "Recommendation Stage": ev.recommendation_stage,
                "Strategy": ev.strategy_source,
                "Telegram Source": ev.telegram_source,
                "Market Condition": ev.market_regime,
                "Entry From": item.entry_zone_low,
                "Entry To": item.entry_zone_high,
                "Stop Loss": item.stop_loss,
                "Target 1": item.target_1,
                "Target 2": item.target_2,
                "Signal Price": ev.signal_price,
                "Latest Close": ev.latest_close,
                "Highest After Signal": ev.highest_after_signal,
                "Lowest After Signal": ev.lowest_after_signal,
                "Actual Return %": ev.actual_return_pct,
                "Max Favorable Move %": ev.max_favorable_move_pct,
                "Max Adverse Move %": ev.max_adverse_move_pct,
                "Days Evaluated": ev.days_evaluated,
                "Status": ev.final_status,
                "Quality": ev.final_quality,
                "Notes": ev.evaluation_notes,
                "Evaluated At": ev.evaluated_at,
            }
        )
    return pd.DataFrame(data)


def _group_accuracy(df: pd.DataFrame, group_col: str) -> pd.DataFrame:
    if df.empty or group_col not in df.columns:
        return pd.DataFrame()
    rows: list[dict[str, Any]] = []
    for key, group in df.groupby(group_col, dropna=False):
        valid = group[~group["Status"].isin(list(NON_ACCURACY_STATUSES))]
        target = valid[valid["Status"].eq(EVAL_TARGET_HIT)]
        returns = pd.to_numeric(valid["Actual Return %"], errors="coerce").dropna()
        rows.append(
            {
                group_col: key if pd.notna(key) else "Unknown",
                "total": len(group),
                "evaluated": len(valid),
                "target_hits": len(target),
                "stop_hits": int(valid["Status"].eq(EVAL_STOP_HIT).sum()),
                "win_rate_pct": round((len(target) / len(valid)) * 100.0, 2) if len(valid) >= 5 else None,
                "accuracy_note": "" if len(valid) >= 5 else "Not reliable yet; evaluated sample size is below 5.",
                "avg_return_pct": round(float(returns.mean()), 2) if not returns.empty else None,
            }
        )
    return pd.DataFrame(rows)


def build_performance_frames(db: Session, *, limit: int = 5000) -> dict[str, pd.DataFrame]:
    stock_rows = _all_evaluations_frame(db, limit=limit)
    summary = summarize_evaluations(
        [
            {
                "symbol": row.get("Stock Symbol"),
                "final_status": row.get("Status"),
                "actual_return_pct": row.get("Actual Return %"),
            }
            for row in stock_rows.to_dict("records")
        ]
    )
    return {
        "summary": pd.DataFrame([{"Metric": key, "Value": value} for key, value in summary.items()]),
        "stock_by_stock": stock_rows,
        "accuracy_by_stage": _group_accuracy(stock_rows, "Recommendation Stage"),
        "accuracy_by_strategy": _group_accuracy(stock_rows, "Strategy"),
        "accuracy_by_telegram_source": _group_accuracy(stock_rows, "Telegram Source"),
        "accuracy_by_market_condition": _group_accuracy(stock_rows, "Market Condition"),
    }


def _rows_for_evaluation_date(df: pd.DataFrame, as_of_date: date) -> pd.DataFrame:
    if df.empty or "Evaluated At" not in df.columns:
        return pd.DataFrame()
    dates = pd.to_datetime(df["Evaluated At"], errors="coerce").dt.date
    return df[dates == as_of_date].copy()


def _summary_records_from_frame(df: pd.DataFrame) -> list[dict[str, Any]]:
    if df.empty:
        return []
    return [
        {
            "symbol": row.get("Stock Symbol"),
            "final_status": row.get("Status"),
            "actual_return_pct": row.get("Actual Return %"),
            "target_hit": row.get("Status") == EVAL_TARGET_HIT,
            "stop_hit": row.get("Status") == EVAL_STOP_HIT,
        }
        for row in df.to_dict("records")
    ]


def write_performance_excel(frames: dict[str, pd.DataFrame], path: str | Path) -> Path:
    from openpyxl.styles import Font, PatternFill

    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    sheet_map = {
        "summary": "Summary",
        "stock_by_stock": "Stock by Stock Comparison",
        "accuracy_by_stage": "Accuracy by Stage",
        "accuracy_by_strategy": "Accuracy by Strategy",
        "accuracy_by_telegram_source": "Accuracy by Telegram Source",
        "accuracy_by_market_condition": "Accuracy by Market Condition",
    }
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for key, sheet_name in sheet_map.items():
            df = frames.get(key)
            if df is None or df.empty:
                df = pd.DataFrame([{"Status": "No data available"}])
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        workbook = writer.book
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        fills = {
            EVAL_TARGET_HIT: PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            EVAL_STOP_HIT: PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            EVAL_EVALUATED: PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
            EVAL_NOT_EVALUATED: PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            EVAL_ENTRY_NOT_REACHED: PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid"),
            EVAL_DATA_MISSING: PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            EVAL_EXPIRED: PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
        }
        for ws in workbook.worksheets:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            headers = {str(cell.value or ""): idx + 1 for idx, cell in enumerate(ws[1])}
            status_col = headers.get("Status")
            if status_col:
                for row_idx in range(2, ws.max_row + 1):
                    value = str(ws.cell(row=row_idx, column=status_col).value or "")
                    fill = fills.get(value)
                    if fill:
                        ws.cell(row=row_idx, column=status_col).fill = fill
            for column_cells in ws.columns:
                values = [str(cell.value or "") for cell in column_cells]
                ws.column_dimensions[column_cells[0].column_letter].width = min(52, max(12, max(len(value) for value in values) + 2))
    return output


def generate_performance_excel(*, as_of_date: date | None = None, db: Session | None = None) -> Path:
    as_of_date = as_of_date or cairo_now().date()

    def _run(active_db: Session) -> Path:
        frames = build_performance_frames(active_db)
        path = REPORT_DIR / f"Recommendation_Performance_{as_of_date:%Y%m%d}_{cairo_now():%H%M%S}.xlsx"
        return write_performance_excel(frames, path)

    if db is not None:
        return _run(db)
    with SessionLocal() as active_db:
        return _run(active_db)


def format_telegram_summary(summary: dict[str, Any], *, as_of_date: date, excel_path: Path | None = None, file_too_large: bool = False) -> str:
    lines = [
        "EGX Recommendation Performance Report",
        f"Date: {as_of_date.isoformat()}",
        "",
        f"Total recommendations: {summary.get('total_recommendations', 0)}",
        f"Evaluated recommendations: {summary.get('evaluated_recommendations', 0)}",
        f"Not evaluated recommendations: {summary.get('not_evaluated_recommendations', 0)}",
        f"Missing data recommendations: {summary.get('missing_data_recommendations', 0)}",
        f"Entry not reached: {summary.get('entry_not_reached_recommendations', 0)}",
        f"Newly evaluated today: {summary.get('newly_evaluated_today', 0)}",
        f"Target hits today: {summary.get('target_hits_today', 0)}",
        f"Stop hits today: {summary.get('stop_hits_today', 0)}",
        f"Open recommendations: {summary.get('open_recommendations', 0)}",
    ]
    if summary.get("win_rate_pct") is None:
        lines.append(f"Win rate: {summary.get('accuracy_note') or 'Not reliable yet.'}")
    else:
        lines.append(f"Win rate: {summary.get('win_rate_pct')}%")
    lines.extend(
        [
            f"Average return %: {summary.get('average_return_pct') if summary.get('average_return_pct') is not None else '-'}",
            f"Best stock today: {summary.get('best_stock_today') or '-'}",
            f"Worst stock today: {summary.get('worst_stock_today') or '-'}",
        ]
    )
    if excel_path:
        lines.append(f"Excel report: {'attached' if not file_too_large else 'too large to attach'}")
        if file_too_large:
            lines.append(f"Saved locally: {excel_path}")
    lines.extend(["", "System remains in audit/paper mode. Live trading is disabled.", f"Risk Note: {RISK_NOTE}"])
    return "\n".join(lines)


def _row_line(row: dict[str, Any]) -> str:
    status = row.get("Status") or EVAL_NOT_EVALUATED
    notes = row.get("Notes") or (
        "Not evaluated yet - waiting for future market data."
        if status == EVAL_NOT_EVALUATED
        else "Entry zone was not reached; no trade should be counted."
        if status == EVAL_ENTRY_NOT_REACHED
        else "-"
    )
    ret = row.get("Actual Return %")
    ret_text = "-" if ret is None or pd.isna(ret) else f"{float(ret):+.2f}%"
    return (
        f"{row.get('Stock Symbol')}\n"
        f"Recommendation: {row.get('Recommendation Stage') or '-'} | Date: {row.get('Recommendation Date') or '-'}\n"
        f"Entry: {row.get('Entry From') or '-'} - {row.get('Entry To') or '-'} | Target: {row.get('Target 1') or '-'} | Stop: {row.get('Stop Loss') or '-'}\n"
        f"Actual: latest close {row.get('Latest Close') or '-'} | High {row.get('Highest After Signal') or '-'} | Low {row.get('Lowest After Signal') or '-'}\n"
        f"Result: {status} | Return: {ret_text} | Quality: {row.get('Quality') or '-'}\n"
        f"Notes: {notes}"
    )


def split_telegram_messages(texts: list[str], *, limit: int = 3600) -> list[str]:
    chunks: list[str] = []
    current = ""
    for text in texts:
        block = text.strip()
        if not block:
            continue
        candidate = f"{current}\n\n{block}".strip() if current else block
        if len(candidate) <= limit:
            current = candidate
            continue
        if current:
            chunks.append(current)
        if len(block) <= limit:
            current = block
        else:
            for idx in range(0, len(block), limit):
                chunks.append(block[idx:idx + limit])
            current = ""
    if current:
        chunks.append(current)
    return chunks


def stock_by_stock_messages(df: pd.DataFrame) -> list[str]:
    if df.empty:
        return ["Stock-by-stock comparison: no recommendation rows are available yet."]
    lines = ["Stock-by-stock Recommendation Review"]
    for row in df.to_dict("records"):
        lines.append(_row_line(row))
    return split_telegram_messages(lines)


def _notification_hash(as_of_date: date) -> str:
    return hashlib.sha256(f"recommendation_performance_report:{as_of_date.isoformat()}".encode("utf-8")).hexdigest()


def _duplicate_sent(db: Session, as_of_date: date) -> bool:
    return db.scalar(select(NotificationLog).where(NotificationLog.notification_hash == _notification_hash(as_of_date))) is not None


def _mark_sent(db: Session, as_of_date: date, *, sent: bool, score: float | None = None) -> None:
    if _duplicate_sent(db, as_of_date):
        return
    db.add(
        NotificationLog(
            notification_hash=_notification_hash(as_of_date),
            symbol="ALL",
            notification_type="recommendation_performance_report",
            recommendation="REPORT",
            score=score,
            source_module="recommendation_performance",
            delivery_status="sent" if sent else "failed",
            cooldown_applied=False,
        )
    )


def send_performance_report_to_telegram(
    *,
    as_of_date: date | None = None,
    force: bool = False,
    db: Session | None = None,
) -> dict[str, Any]:
    from app.services.telegram_bot import send_private_documents_sync, send_private_message_sync

    as_of_date = as_of_date or cairo_now().date()
    settings = get_settings()

    def _run(active_db: Session) -> dict[str, Any]:
        if _duplicate_sent(active_db, as_of_date) and not force:
            return {"status": "duplicate_skipped", "sent": False}
        frames = build_performance_frames(active_db)
        summary = {row["Metric"]: row["Value"] for row in frames["summary"].to_dict("records")}
        daily_rows = _rows_for_evaluation_date(frames["stock_by_stock"], as_of_date)
        daily_summary = summarize_evaluations(_summary_records_from_frame(daily_rows))
        summary["newly_evaluated_today"] = daily_summary["evaluated_recommendations"]
        summary["target_hits_today"] = daily_summary["target_hits_today"]
        summary["stop_hits_today"] = daily_summary["stop_hits_today"]
        if daily_summary.get("best_stock_today") != "-":
            summary["best_stock_today"] = daily_summary["best_stock_today"]
        if daily_summary.get("worst_stock_today") != "-":
            summary["worst_stock_today"] = daily_summary["worst_stock_today"]
        excel_path = generate_performance_excel(as_of_date=as_of_date, db=active_db)
        file_too_large = excel_path.exists() and excel_path.stat().st_size > 45 * 1024 * 1024
        summary_text = format_telegram_summary(summary, as_of_date=as_of_date, excel_path=excel_path, file_too_large=file_too_large)
        docs = [] if file_too_large else [excel_path]
        doc_result = send_private_documents_sync(summary_text, docs, settings=settings)
        chunks = stock_by_stock_messages(frames["stock_by_stock"])
        sent_chunks = 0
        for chunk in chunks:
            send_private_message_sync(chunk, settings=settings)
            sent_chunks += 1
        sent = bool(doc_result.get("sent_messages") or sent_chunks)
        _mark_sent(active_db, as_of_date, sent=sent, score=_safe_float(summary.get("average_return_pct")))
        active_db.commit()
        return {
            "status": "sent" if sent else "telegram_failed",
            "sent": sent,
            "summary_messages": doc_result.get("sent_messages", 0),
            "sent_documents": doc_result.get("sent_documents", 0),
            "stock_by_stock_chunks": sent_chunks,
            "excel_path": str(excel_path),
            "file_too_large": file_too_large,
            "errors": doc_result.get("errors", []),
        }

    if db is not None:
        return _run(db)
    with sqlite_write_lock():
        with SessionLocal() as active_db:
            return _run(active_db)


def run_re_evaluation_report(
    *,
    as_of_date: date | None = None,
    send_telegram: bool = False,
    force_send: bool = False,
    db: Session | None = None,
) -> dict[str, Any]:
    as_of_date = as_of_date or cairo_now().date()

    def _run(active_db: Session) -> dict[str, Any]:
        evaluation = run_daily_re_evaluation(as_of_date=as_of_date, db=active_db)
        excel_path = generate_performance_excel(as_of_date=as_of_date, db=active_db)
        telegram = None
        if send_telegram:
            telegram = send_performance_report_to_telegram(as_of_date=as_of_date, force=force_send, db=active_db)
        return {"evaluation": evaluation, "excel_path": str(excel_path), "telegram": telegram}

    if db is not None:
        return _run(db)
    init_db(seed=True)
    with sqlite_write_lock():
        with SessionLocal() as active_db:
            return _run(active_db)


def _parse_date(value: str | None) -> date:
    if not value or value.lower() == "today":
        return cairo_now().date()
    return date.fromisoformat(value)


def _cli() -> None:
    parser = argparse.ArgumentParser(description="Re-evaluate old recommendation results and build performance reports.")
    parser.add_argument("--date", default="today", help="today or YYYY-MM-DD")
    parser.add_argument("--send-telegram", action="store_true", help="Send the performance report and Excel to Telegram.")
    parser.add_argument("--force-send", action="store_true", help="Bypass duplicate Telegram report protection.")
    parser.add_argument("--excel-only", action="store_true", help="Only generate the Excel performance report.")
    parser.add_argument("--evaluate-only", action="store_true", help="Only update stored recommendation evaluations.")
    parser.add_argument("--include-today", action="store_true", help="Include same-day recommendation rows, but still never uses same candle as post-signal data.")
    args = parser.parse_args()
    day = _parse_date(args.date)
    if args.excel_only:
        path = generate_performance_excel(as_of_date=day)
        print(json.dumps({"status": "success", "excel_path": str(path)}, indent=2))
        return
    if args.evaluate_only:
        result = run_daily_re_evaluation(as_of_date=day, include_today=args.include_today)
        print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
        return
    result = run_re_evaluation_report(as_of_date=day, send_telegram=args.send_telegram, force_send=args.force_send)
    print(json.dumps(result, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    _cli()
