from __future__ import annotations

import argparse
import hashlib
import json
import logging
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.chart import BarChart, Reference
from sqlalchemy import delete, func, select
from sqlalchemy.orm import Session

from app.config import REPORT_TIMEZONE, RISK_NOTE, get_settings
from app.data.market_data import get_ohlcv
from app.database import SessionLocal, init_db, sqlite_write_lock
from app.models import (
    EndOfDayReviewItem,
    EndOfDayReviewReport,
    FinancialSignal,
    NotificationLog,
    Stock,
    TelegramMessageSymbol,
)
from app.services.daily_loss_audit import (
    EVAL_DATA_MISSING,
    EVAL_ENTRY_NOT_REACHED,
    EVAL_NOT_EVALUATED,
    EVAL_STOP_HIT,
    EVAL_TARGET_HIT,
    build_daily_loss_audit,
)
from app.services.daily_stock_report import build_report_items
from app.services.market_daily_evaluation import evaluate_daily_market
from app.technical.indicators import add_indicators
from app.technical.support_resistance import breakout_state, support_resistance


logger = logging.getLogger(__name__)
CAIRO_TZ = ZoneInfo(REPORT_TIMEZONE)
PROJECT_ROOT = Path(__file__).resolve().parents[2]
REPORT_DIR = PROJECT_ROOT / "reports" / "end_of_day"
REPORT_DIR.mkdir(parents=True, exist_ok=True)

NON_ACCURACY_STATUSES = {EVAL_NOT_EVALUATED, EVAL_DATA_MISSING, EVAL_ENTRY_NOT_REACHED}
SELECTION_THRESHOLD = 70.0
MIN_VALUE_TRADED = 500_000.0


def cairo_now() -> datetime:
    return datetime.now(CAIRO_TZ)


def _parse_date(value: str | None) -> date:
    if not value or str(value).lower() == "today":
        return cairo_now().date()
    return date.fromisoformat(str(value))


def _day_bounds(day: date) -> tuple[datetime, datetime]:
    start = datetime(day.year, day.month, day.day)
    return start, start + timedelta(days=1)


def _safe_float(value: Any) -> float | None:
    if value is None:
        return None
    try:
        number = float(value)
    except Exception:
        return None
    if pd.isna(number):
        return None
    return number


def _round(value: Any, digits: int = 2) -> float | None:
    number = _safe_float(value)
    return round(number, digits) if number is not None else None


def _score(value: Any, default: float = 50.0) -> float:
    number = _safe_float(value)
    if number is None:
        return default
    return max(0.0, min(100.0, number))


def _latest_financial_score(db: Session, symbol: str) -> float:
    row = db.scalar(
        select(FinancialSignal)
        .where(FinancialSignal.symbol == symbol, FinancialSignal.financial_score.is_not(None))
        .order_by(FinancialSignal.signal_date.desc(), FinancialSignal.id.desc())
    )
    return _score(row.financial_score if row else None, 50.0)


def _telegram_mentions(db: Session, symbol: str, day: date) -> dict[str, Any]:
    start, end = _day_bounds(day)
    rows = db.execute(
        select(
            func.count(TelegramMessageSymbol.id).label("mentions"),
            func.avg(TelegramMessageSymbol.confidence).label("avg_confidence"),
        )
        .where(
            TelegramMessageSymbol.symbol == symbol,
            TelegramMessageSymbol.created_at >= start,
            TelegramMessageSymbol.created_at < end,
        )
    ).mappings().one()
    return {
        "mentions": int(rows["mentions"] or 0),
        "avg_confidence": _round(rows["avg_confidence"]),
    }


def daily_move_from_ohlcv(df: pd.DataFrame, target_date: date) -> dict[str, Any]:
    if df is None or df.empty or "datetime" not in df.columns:
        return {"status": "DATA_MISSING", "reason": "No OHLCV data available."}
    frame = df.copy().reset_index(drop=True)
    frame["datetime"] = pd.to_datetime(frame["datetime"], errors="coerce")
    frame = frame.dropna(subset=["datetime"]).sort_values("datetime")
    day_rows = frame[frame["datetime"].dt.date == target_date]
    if day_rows.empty:
        return {"status": "DATA_MISSING", "reason": f"No candle for {target_date.isoformat()}."}
    row = day_rows.iloc[-1]
    prev = frame[frame["datetime"].dt.date < target_date].tail(20)
    prev_close = _safe_float(prev.iloc[-1].get("close")) if not prev.empty else None
    open_price = _safe_float(row.get("open"))
    close = _safe_float(row.get("close"))
    high = _safe_float(row.get("high"))
    low = _safe_float(row.get("low"))
    volume = _safe_float(row.get("volume")) or 0.0
    base = prev_close or open_price
    return_pct = ((close - base) / base * 100.0) if close is not None and base not in {None, 0} else None
    prior_volume = float(prev["volume"].fillna(0).mean() or 0.0) if not prev.empty and "volume" in prev.columns else 0.0
    volume_change = ((volume - prior_volume) / prior_volume * 100.0) if prior_volume > 0 else None
    value_traded = close * volume if close is not None else None
    enriched = frame[frame["datetime"].dt.date <= target_date].tail(120).copy()
    breakout = "UNKNOWN"
    support = resistance = None
    momentum_acceleration = None
    try:
        if len(enriched) >= 30:
            indicators = add_indicators(enriched)
            breakout = breakout_state(indicators)
            sr = support_resistance(indicators)
            support = sr.get("support")
            resistance = sr.get("resistance")
            last = indicators.iloc[-1]
            prev_last = indicators.iloc[-2]
            momentum_acceleration = _round((last.get("macd_hist") or 0) - (prev_last.get("macd_hist") or 0), 4)
    except Exception as exc:
        logger.debug("Daily move indicator enrichment failed: %s", exc)
    return {
        "status": "OK",
        "open": _round(open_price, 4),
        "high": _round(high, 4),
        "low": _round(low, 4),
        "close": _round(close, 4),
        "prev_close": _round(prev_close, 4),
        "return_pct": _round(return_pct),
        "volume": _round(volume, 0),
        "volume_change_pct": _round(volume_change),
        "value_traded": _round(value_traded, 0),
        "breakout_signal": breakout,
        "support": _round(support, 4),
        "resistance": _round(resistance, 4),
        "momentum_acceleration": momentum_acceleration,
    }


def _filters_from_scores(row: dict[str, Any], market: dict[str, Any], move: dict[str, Any]) -> tuple[list[str], list[str]]:
    passed: list[str] = []
    failed: list[str] = []
    checks = [
        ("technical_score", row.get("technical_score"), 70),
        ("strategy_score", row.get("strategy_score"), 65),
        ("news_score", row.get("news_score"), 40),
        ("telegram_score", row.get("telegram_score"), 45),
        ("risk_liquidity_score", row.get("risk_liquidity_score"), 60),
        ("final_score", row.get("final_score"), SELECTION_THRESHOLD),
    ]
    for name, value, threshold in checks:
        score = _score(value)
        (passed if score >= threshold else failed).append(f"{name}>={threshold}")
    rr = _safe_float(row.get("risk_reward"))
    (passed if rr is not None and rr >= 1.8 else failed).append("risk_reward>=1.8")
    entry_valid = bool((row.get("details") or {}).get("entry_zone", {}).get("valid"))
    (passed if entry_valid else failed).append("entry_zone_valid")
    permission = str(market.get("trade_permission") or "DATA_INSUFFICIENT")
    (passed if permission in {"TRADE_ALLOWED", "WATCH_ONLY"} else failed).append(f"market_permission={permission}")
    value_traded = _safe_float(move.get("value_traded"))
    (passed if value_traded is not None and value_traded >= MIN_VALUE_TRADED else failed).append("liquidity_value>=500k")
    return passed, failed


def classify_why_not_selected(row: dict[str, Any], market: dict[str, Any], move: dict[str, Any]) -> str:
    if move.get("status") == "DATA_MISSING":
        return "DATA_MISSING"
    if (_safe_float(move.get("value_traded")) or 0) < MIN_VALUE_TRADED or _score(row.get("risk_liquidity_score")) < 40:
        return "LOW_LIQUIDITY"
    if str(market.get("trade_permission") or "") in {"BUY_BLOCKED", "NO_TRADING", "SELL_ONLY", "DATA_INSUFFICIENT"}:
        return "MARKET_REGIME_BLOCKED"
    if _score(row.get("news_score")) < 40:
        return "NEGATIVE_NEWS"
    if _safe_float(row.get("risk_reward")) is None or (_safe_float(row.get("risk_reward")) or 0) < 1.8:
        return "RISK_REWARD_FAILED"
    if _score(row.get("technical_score")) < 60:
        return "WEAK_TECHNICAL_SCORE"
    if _score(row.get("strategy_score")) < 50 and _score(row.get("technical_score")) >= 70:
        return "STRATEGY_NOT_COVERED"
    if _score(row.get("telegram_score")) < 45 and int((row.get("telegram_mentions") or 0)) > 0:
        return "LOW_TELEGRAM_CONFIDENCE"
    if (_safe_float(move.get("return_pct")) or 0) >= 5 and (_safe_float(move.get("volume_change_pct")) or 0) >= 50:
        return "LATE_BREAKOUT"
    if _score(row.get("final_score")) >= SELECTION_THRESHOLD - 5:
        return "FILTER_TOO_STRICT"
    return "UNKNOWN_REASON"


def _suggested_fix(reason: str) -> str:
    mapping = {
        "DATA_MISSING": "Refresh TradingView/OHLCV data before the close and mark missing data clearly.",
        "LOW_LIQUIDITY": "Keep liquidity gate active; only relax after validating spread and traded value.",
        "SPREAD_TOO_WIDE": "Add bid/ask spread feed before allowing intraday entry.",
        "WEAK_TECHNICAL_SCORE": "Add momentum acceleration and breakout re-test logic for fast movers.",
        "WEAK_FINANCIAL_SCORE": "Do not loosen financial filters without reviewing latest statements.",
        "NEGATIVE_NEWS": "Keep news risk as a blocker unless technical confirmation is exceptional.",
        "LOW_TELEGRAM_CONFIDENCE": "Improve channel accuracy weighting and OCR confidence before raising Telegram weight.",
        "LATE_BREAKOUT": "Add intraday re-scans around 11:30 AM and 1:30 PM for volume breakouts.",
        "FILTER_TOO_STRICT": "Review threshold margins; consider conditional watchlist instead of direct BUY.",
        "STRATEGY_NOT_COVERED": "Add a volume/momentum breakout strategy branch for gap-and-go moves.",
        "MARKET_REGIME_BLOCKED": "Only override market filter manually after daily market review.",
        "RISK_REWARD_FAILED": "Improve entry-zone calculation and require realistic pullback/retest.",
        "DUPLICATE_SIGNAL_BLOCKED": "Keep duplicate blocking but roll repeated confirmed signals into watchlist score.",
        "UNKNOWN_REASON": "Inspect component scores and add a more specific rejection rule.",
    }
    return mapping.get(reason, mapping["UNKNOWN_REASON"])


def _why_text(reason: str, failed_filters: list[str]) -> str:
    filter_text = ", ".join(failed_filters[:5]) if failed_filters else "No hard failed filter recorded."
    return f"{reason}: {filter_text}"


def _stock_name(stock: Stock | None, symbol: str) -> str:
    return (stock.name or stock.name_en or stock.name_ar) if stock else symbol


def _market_score(market: dict[str, Any]) -> float:
    return _score(market.get("market_score"), 50.0)


def _result_rows_from_audit(audit: dict[str, Any], stock_map: dict[str, Stock]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for item in audit.get("items") or []:
        symbol = str(item.get("symbol") or "").upper()
        stock = stock_map.get(symbol)
        rows.append(
            {
                "Report Date": audit.get("audit_date"),
                "Recommendation Date": item.get("recommendation_datetime"),
                "Stock Symbol": symbol,
                "Stock Name": _stock_name(stock, symbol),
                "Sector": stock.sector if stock else None,
                "Recommendation Stage": item.get("recommended_signal"),
                "Entry From": item.get("entry_zone_low"),
                "Entry To": item.get("entry_zone_high"),
                "Stop Loss": item.get("stop_loss"),
                "Target 1": item.get("target_1"),
                "Target 2": item.get("target_2"),
                "Actual Close": item.get("latest_close"),
                "Highest After Recommendation": item.get("highest_price_after_signal"),
                "Lowest After Recommendation": item.get("lowest_price_after_signal"),
                "Actual Return %": item.get("actual_return"),
                "Target Hit": bool(item.get("details", {}).get("target_1_hit") or item.get("details", {}).get("target_2_hit") or item.get("details", {}).get("target_3_hit")),
                "Stop Hit": bool(item.get("details", {}).get("stop_loss_hit")),
                "Final Result": item.get("evaluation_status"),
                "Quality Rating": item.get("final_quality"),
                "Reason": item.get("root_cause"),
                "Technical Score": item.get("technical_score"),
                "Financial Score": _score(None, 50.0),
                "News Score": item.get("news_score"),
                "Telegram Score": item.get("telegram_score"),
                "Strategy Score": item.get("strategy_score"),
                "Liquidity Score": item.get("risk_liquidity_score"),
                "Backtest Score": item.get("backtest_score"),
                "Evaluation Quality": item.get("evaluation_quality"),
            }
        )
    return rows


def _summary_from_results(results: list[dict[str, Any]], missed: list[dict[str, Any]]) -> dict[str, Any]:
    valid = [row for row in results if row.get("Final Result") not in NON_ACCURACY_STATUSES]
    targets = [row for row in valid if row.get("Final Result") == EVAL_TARGET_HIT or row.get("Target Hit")]
    stops = [row for row in valid if row.get("Final Result") == EVAL_STOP_HIT or row.get("Stop Hit")]
    entry_not = [row for row in results if row.get("Final Result") == EVAL_ENTRY_NOT_REACHED]
    returns = [_safe_float(row.get("Actual Return %")) for row in valid]
    returns = [value for value in returns if value is not None]
    best = max(valid, key=lambda row: _safe_float(row.get("Actual Return %")) if _safe_float(row.get("Actual Return %")) is not None else -999999, default=None)
    worst = min(valid, key=lambda row: _safe_float(row.get("Actual Return %")) if _safe_float(row.get("Actual Return %")) is not None else 999999, default=None)
    return {
        "total_recommendations": len(results),
        "evaluated_recommendations": len(valid),
        "not_evaluated": sum(1 for row in results if row.get("Final Result") == EVAL_NOT_EVALUATED),
        "data_missing": sum(1 for row in results if row.get("Final Result") == EVAL_DATA_MISSING),
        "entry_not_reached": len(entry_not),
        "target_hits": len(targets),
        "stop_hits": len(stops),
        "open_trades": sum(1 for row in valid if row.get("Final Result") == "EVALUATED"),
        "win_rate_pct": round(len(targets) / len(valid) * 100.0, 2) if len(valid) >= 5 else None,
        "accuracy_note": "" if len(valid) >= 5 else "Accuracy is not reliable yet because evaluated sample size is below 5.",
        "average_return_pct": round(sum(returns) / len(returns), 2) if returns else None,
        "best_recommendation": f"{best.get('Stock Symbol')} ({best.get('Actual Return %')}%)" if best else "-",
        "worst_recommendation": f"{worst.get('Stock Symbol')} ({worst.get('Actual Return %')}%)" if worst else "-",
        "top_missed_opportunity": f"{missed[0].get('Stock Symbol')} ({missed[0].get('Today Return %')}%)" if missed else "-",
    }


def _strategy_suggestions(results: list[dict[str, Any]], missed: list[dict[str, Any]], why_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    reasons = pd.Series([row.get("Why Not Selected Code") for row in missed if row.get("Why Not Selected Code")]).value_counts()
    suggestions: list[dict[str, Any]] = []
    if not reasons.empty:
        top_reason = str(reasons.index[0])
        suggestions.append({"Area": top_reason, "Suggestion": _suggested_fix(top_reason), "Auto Applied": "No"})
    if any(row.get("Why Not Selected Code") == "LATE_BREAKOUT" for row in missed):
        suggestions.append({"Area": "Intraday Scans", "Suggestion": "Add scans around 11:30 AM and 1:30 PM to catch late breakouts.", "Auto Applied": "No"})
    if any((_safe_float(row.get("Telegram Score")) or 0) >= 75 and (_safe_float(row.get("Technical Score")) or 0) < 60 for row in results):
        suggestions.append({"Area": "Telegram Weight", "Suggestion": "Telegram was strong while technical confirmation was weak; keep Telegram as supporting evidence, not a standalone trigger.", "Auto Applied": "No"})
    if any((_safe_float(row.get("Liquidity Score")) or 0) < 50 for row in why_rows):
        suggestions.append({"Area": "Liquidity", "Suggestion": "Flag low-liquidity names more visibly and keep them out of BUY alerts.", "Auto Applied": "No"})
    if not suggestions:
        suggestions.append({"Area": "General", "Suggestion": "No automatic weight change. Continue collecting more evaluated rows before tuning strategy weights.", "Auto Applied": "No"})
    return suggestions


def build_end_of_day_review(
    *,
    target_date: date | None = None,
    persist: bool = False,
    db: Session | None = None,
    top_movers_limit: int = 10,
) -> dict[str, Any]:
    day = target_date or cairo_now().date()

    def _run(active_db: Session) -> dict[str, Any]:
        stocks = active_db.scalars(select(Stock).where(Stock.is_active.is_(True)).order_by(Stock.symbol.asc())).all()
        stock_map = {stock.symbol.upper(): stock for stock in stocks}
        market = evaluate_daily_market(active_db, target_date=day, persist=False)
        audit = build_daily_loss_audit(target_date=day, persist=False, db=active_db)
        recommendation_results = _result_rows_from_audit(audit, stock_map)
        recommended_symbols = {row["Stock Symbol"] for row in recommendation_results}
        analysis_rows = build_report_items(active_db, top_n=max(len(stocks), 1), now=datetime(day.year, day.month, day.day, 15, 0, 0))
        analysis_by_symbol = {str(row.get("symbol") or "").upper(): row for row in analysis_rows}

        top_movers: list[dict[str, Any]] = []
        why_rows: list[dict[str, Any]] = []
        score_rows: list[dict[str, Any]] = []
        data_quality: list[dict[str, Any]] = []
        telegram_vs_actual: list[dict[str, Any]] = []
        technical_vs_actual: list[dict[str, Any]] = []
        financial_vs_actual: list[dict[str, Any]] = []
        news_vs_actual: list[dict[str, Any]] = []

        for stock in stocks:
            symbol = stock.symbol.upper()
            df = get_ohlcv(active_db, symbol, timeframe="1D", limit=260)
            move = daily_move_from_ohlcv(df, day)
            row = analysis_by_symbol.get(symbol) or {
                "symbol": symbol,
                "company_name": _stock_name(stock, symbol),
                "technical_score": 50,
                "telegram_score": 50,
                "strategy_score": 50,
                "news_score": 50,
                "risk_liquidity_score": 50,
                "final_score": 50,
                "risk_reward": None,
                "details": {},
            }
            telegram_meta = _telegram_mentions(active_db, symbol, day)
            financial_score = _latest_financial_score(active_db, symbol)
            passed, failed = _filters_from_scores(row, market, move)
            reason_code = classify_why_not_selected({**row, **telegram_meta}, market, move)
            selected = symbol in recommended_symbols
            should_have = (not selected) and move.get("status") == "OK" and (_safe_float(move.get("return_pct")) or 0) >= 3 and _score(row.get("technical_score")) >= 60
            why = {
                "Review Date": day.isoformat(),
                "Stock Symbol": symbol,
                "Stock Name": _stock_name(stock, symbol),
                "Sector": stock.sector,
                "Selected Today": selected,
                "Should Have Been Selected": should_have,
                "Why Not Selected Code": "" if selected else reason_code,
                "Why Not Selected": "Selected in daily recommendations." if selected else _why_text(reason_code, failed),
                "Suggested Fix": "" if selected else _suggested_fix(reason_code),
                "Technical Score": _round(row.get("technical_score")),
                "Financial Score": financial_score,
                "News Score": _round(row.get("news_score")),
                "Telegram Score": _round(row.get("telegram_score")),
                "Strategy Score": _round(row.get("strategy_score")),
                "Liquidity Score": _round(row.get("risk_liquidity_score")),
                "Risk/Reward Score": _round((_safe_float(row.get("risk_reward")) or 0) * 40 if row.get("risk_reward") else 0),
                "Market Regime Score": _market_score(market),
                "Final Score": _round(row.get("final_score")),
                "Selection Threshold": SELECTION_THRESHOLD,
                "Passed Filters": ", ".join(passed),
                "Failed Filters": ", ".join(failed),
                "Today Return %": move.get("return_pct"),
                "Volume Change %": move.get("volume_change_pct"),
                "Value Traded": move.get("value_traded"),
                "Breakout Signal": move.get("breakout_signal"),
                "Telegram Mentions": telegram_meta["mentions"],
                "Telegram Avg Confidence": telegram_meta["avg_confidence"],
            }
            why_rows.append(why)
            score_rows.append(why)
            data_quality.append({"Stock Symbol": symbol, "Status": move.get("status"), "Reason": move.get("reason", ""), "Rows": len(df) if df is not None else 0})
            if move.get("status") == "OK":
                top_movers.append({**why, "Status": "TOP_MOVER" if symbol in recommended_symbols else "MISSED_OPPORTUNITY"})
            telegram_vs_actual.append({"Stock Symbol": symbol, "Telegram Score": why["Telegram Score"], "Actual Return %": why["Today Return %"], "Selected": selected})
            technical_vs_actual.append({"Stock Symbol": symbol, "Technical Score": why["Technical Score"], "Actual Return %": why["Today Return %"], "Selected": selected})
            financial_vs_actual.append({"Stock Symbol": symbol, "Financial Score": why["Financial Score"], "Actual Return %": why["Today Return %"], "Selected": selected})
            news_vs_actual.append({"Stock Symbol": symbol, "News Score": why["News Score"], "Actual Return %": why["Today Return %"], "Selected": selected})

        top_movers = sorted(top_movers, key=lambda item: _safe_float(item.get("Today Return %")) or -999999, reverse=True)
        missed = [
            row
            for row in top_movers
            if not row.get("Selected Today") and (_safe_float(row.get("Today Return %")) or 0.0) >= 3.0
        ][:top_movers_limit]
        almost = [row for row in why_rows if not row.get("Selected Today") and (_safe_float(row.get("Final Score")) or 0) >= SELECTION_THRESHOLD - 5]
        strong_volume = [row for row in why_rows if (_safe_float(row.get("Volume Change %")) or 0) >= 50 and not row.get("Selected Today")]
        tomorrow_watchlist = sorted({row["Stock Symbol"]: row for row in (almost + strong_volume + missed)}.values(), key=lambda row: (_safe_float(row.get("Final Score")) or 0), reverse=True)[:20]
        for row in tomorrow_watchlist:
            row["Watchlist Reason"] = "Almost passed, strong volume, or missed top mover. Review before tomorrow's open."
        suggestions = _strategy_suggestions(recommendation_results, missed, why_rows)
        summary = _summary_from_results(recommendation_results, missed)
        payload = {
            "review_date": day.isoformat(),
            "market_evaluation": market,
            "summary": summary,
            "daily_prediction_review": pd.DataFrame([summary]),
            "recommendation_results": pd.DataFrame(recommendation_results),
            "missed_opportunities": pd.DataFrame(missed),
            "why_not_selected": pd.DataFrame(why_rows),
            "top_movers_analysis": pd.DataFrame(top_movers[:30]),
            "strategy_improvement_suggestions": pd.DataFrame(suggestions),
            "score_breakdown": pd.DataFrame(score_rows),
            "data_quality_issues": pd.DataFrame(data_quality),
            "telegram_vs_actual": pd.DataFrame(telegram_vs_actual),
            "technical_vs_actual": pd.DataFrame(technical_vs_actual),
            "financial_vs_actual": pd.DataFrame(financial_vs_actual),
            "news_vs_actual": pd.DataFrame(news_vs_actual),
            "tomorrow_watchlist": pd.DataFrame(tomorrow_watchlist),
            "risk_note": RISK_NOTE,
        }
        try:
            from app.services.learning_system import build_learning_payload, format_learning_telegram_block

            learning_payload = build_learning_payload(
                active_db,
                target_date=day,
                missed_df=payload["missed_opportunities"],
                persist=persist,
            )
            payload.update(
                {
                    "source_accuracy": learning_payload.get("source_accuracy"),
                    "risk_expectancy": learning_payload.get("risk_expectancy"),
                    "walk_forward_periods": learning_payload.get("walk_forward_periods"),
                    "walk_forward_summary": learning_payload.get("walk_forward_summary"),
                    "intraday_scan": learning_payload.get("intraday_scan"),
                    "pump_risk_monitor": learning_payload.get("pump_risk_monitor"),
                    "recommendation_quality": learning_payload.get("recommendation_quality"),
                    "missed_opportunity_diagnosis": learning_payload.get("missed_opportunity_diagnosis"),
                    "decision_snapshots": learning_payload.get("decision_snapshots"),
                    "learning_report": learning_payload.get("strategy_learning"),
                    "learning_telegram_block": format_learning_telegram_block(learning_payload),
                }
            )
        except Exception as exc:
            logger.exception("Learning payload failed for end-of-day review.")
            payload["learning_error"] = str(exc)
        if persist:
            report = _persist_review(active_db, payload)
            payload["report_id"] = report.id
        return payload

    if db is not None:
        return _run(db)
    init_db(seed=True)
    with sqlite_write_lock():
        with SessionLocal() as active_db:
            return _run(active_db)


def _persist_review(db: Session, payload: dict[str, Any]) -> EndOfDayReviewReport:
    review_date = date.fromisoformat(payload["review_date"])
    start, end = _day_bounds(review_date)
    existing = db.scalar(select(EndOfDayReviewReport).where(EndOfDayReviewReport.review_date >= start, EndOfDayReviewReport.review_date < end))
    if existing:
        db.execute(delete(EndOfDayReviewItem).where(EndOfDayReviewItem.report_id == existing.id))
        report = existing
    else:
        report = EndOfDayReviewReport(review_date=start)
        db.add(report)
        db.flush()
    report.status = "created"
    report.summary_json = payload.get("summary")
    report.suggestions_json = payload.get("strategy_improvement_suggestions", pd.DataFrame()).to_dict("records")
    for key, row_type in [
        ("recommendation_results", "RECOMMENDATION_RESULT"),
        ("missed_opportunities", "MISSED_OPPORTUNITY"),
        ("why_not_selected", "WHY_NOT_SELECTED"),
        ("tomorrow_watchlist", "TOMORROW_WATCHLIST"),
    ]:
        df = payload.get(key)
        if not isinstance(df, pd.DataFrame) or df.empty:
            continue
        for row in df.to_dict("records"):
            db.add(
                EndOfDayReviewItem(
                    report_id=report.id,
                    review_date=start,
                    row_type=row_type,
                    symbol=str(row.get("Stock Symbol") or row.get("symbol") or ""),
                    stock_name=row.get("Stock Name"),
                    sector=row.get("Sector"),
                    recommendation_stage=row.get("Recommendation Stage"),
                    classification=row.get("Why Not Selected Code") or row.get("Status"),
                    final_status=row.get("Final Result"),
                    final_quality=row.get("Quality Rating"),
                    actual_return_pct=_safe_float(row.get("Actual Return %") or row.get("Today Return %")),
                    volume_change_pct=_safe_float(row.get("Volume Change %")),
                    value_traded=_safe_float(row.get("Value Traded")),
                    technical_score=_safe_float(row.get("Technical Score")),
                    financial_score=_safe_float(row.get("Financial Score")),
                    news_score=_safe_float(row.get("News Score")),
                    telegram_score=_safe_float(row.get("Telegram Score")),
                    liquidity_score=_safe_float(row.get("Liquidity Score")),
                    risk_reward_score=_safe_float(row.get("Risk/Reward Score")),
                    market_regime_score=_safe_float(row.get("Market Regime Score")),
                    final_score=_safe_float(row.get("Final Score")),
                    selection_threshold=_safe_float(row.get("Selection Threshold")),
                    passed_filters_json=[part.strip() for part in str(row.get("Passed Filters") or "").split(",") if part.strip()],
                    failed_filters_json=[part.strip() for part in str(row.get("Failed Filters") or "").split(",") if part.strip()],
                    reason=row.get("Why Not Selected") or row.get("Reason"),
                    suggested_fix=row.get("Suggested Fix"),
                    details_json=row,
                )
            )
    db.commit()
    return report


def write_end_of_day_excel(payload: dict[str, Any], path: str | Path | None = None) -> Path:
    day = date.fromisoformat(payload["review_date"])
    output = Path(path) if path else REPORT_DIR / f"EGX_End_of_Day_Review_{day:%Y%m%d}_{cairo_now():%H%M%S}.xlsx"
    output.parent.mkdir(parents=True, exist_ok=True)
    sheet_map = {
        "daily_prediction_review": "Daily Prediction Review",
        "recommendation_results": "Recommendation Results",
        "missed_opportunities": "Missed Opportunities",
        "why_not_selected": "Why Not Selected",
        "top_movers_analysis": "Top Movers Analysis",
        "strategy_improvement_suggestions": "Strategy Improvement Suggestion",
        "score_breakdown": "Score Breakdown",
        "data_quality_issues": "Data Quality Issues",
        "telegram_vs_actual": "Telegram vs Actual",
        "technical_vs_actual": "Technical vs Actual",
        "financial_vs_actual": "Financial vs Actual",
        "news_vs_actual": "News vs Actual",
        "tomorrow_watchlist": "Tomorrow Watchlist",
        "source_accuracy": "Source Accuracy",
        "risk_expectancy": "Risk Expectancy",
        "walk_forward_periods": "Walk Forward Testing",
        "intraday_scan": "Intraday Scanner",
        "pump_risk_monitor": "Pump Risk Monitor",
        "recommendation_quality": "Recommendation Quality",
        "missed_opportunity_diagnosis": "Missed Deep Diagnosis",
        "decision_snapshots": "Decision Snapshots",
    }
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for key, sheet_name in sheet_map.items():
            df = payload.get(key)
            if not isinstance(df, pd.DataFrame) or df.empty:
                df = pd.DataFrame([{"Status": "No data available"}])
            df.to_excel(writer, index=False, sheet_name=sheet_name)
        wb = writer.book
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        status_fills = {
            EVAL_TARGET_HIT: PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            EVAL_STOP_HIT: PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            EVAL_ENTRY_NOT_REACHED: PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid"),
            EVAL_NOT_EVALUATED: PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            EVAL_DATA_MISSING: PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
        }
        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            headers = {str(cell.value or ""): idx + 1 for idx, cell in enumerate(ws[1])}
            status_col = headers.get("Final Result") or headers.get("Status")
            if status_col:
                for row_idx in range(2, ws.max_row + 1):
                    fill = status_fills.get(str(ws.cell(row=row_idx, column=status_col).value or ""))
                    if fill:
                        ws.cell(row=row_idx, column=status_col).fill = fill
            for column_cells in ws.columns:
                values = [str(cell.value or "") for cell in column_cells]
                ws.column_dimensions[column_cells[0].column_letter].width = min(54, max(12, max(len(value) for value in values) + 2))
        if "Top Movers Analysis" in wb.sheetnames:
            ws = wb["Top Movers Analysis"]
            headers = {str(cell.value or ""): idx + 1 for idx, cell in enumerate(ws[1])}
            ret_col = headers.get("Today Return %")
            symbol_col = headers.get("Stock Symbol")
            if ret_col and symbol_col and ws.max_row >= 3:
                chart = BarChart()
                chart.title = "Top Movers Return %"
                chart.y_axis.title = "Return %"
                chart.x_axis.title = "Stock"
                chart.add_data(Reference(ws, min_col=ret_col, min_row=1, max_row=min(ws.max_row, 11)), titles_from_data=True)
                chart.set_categories(Reference(ws, min_col=symbol_col, min_row=2, max_row=min(ws.max_row, 11)))
                ws.add_chart(chart, "N2")
    return output


def format_end_of_day_telegram(payload: dict[str, Any], *, excel_path: Path | None = None, dry_run: bool = True) -> list[str]:
    summary = payload.get("summary") or {}
    missed = payload.get("missed_opportunities")
    results = payload.get("recommendation_results")
    suggestions = payload.get("strategy_improvement_suggestions")
    lines = [
        "EGX End-of-Day Prediction Review",
        f"Date: {payload.get('review_date')}",
        "Mode: AUDIT/PAPER ONLY - live trading disabled.",
        "",
        f"Total recommendations: {summary.get('total_recommendations', 0)}",
        f"Evaluated recommendations: {summary.get('evaluated_recommendations', 0)}",
        f"Target hits: {summary.get('target_hits', 0)}",
        f"Stop hits: {summary.get('stop_hits', 0)}",
        f"Entry not reached: {summary.get('entry_not_reached', 0)}",
        f"Open trades: {summary.get('open_trades', 0)}",
    ]
    if summary.get("win_rate_pct") is None:
        lines.append(f"Win rate: {summary.get('accuracy_note')}")
    else:
        lines.append(f"Win rate: {summary.get('win_rate_pct')}%")
    lines.extend(
        [
            f"Average return: {summary.get('average_return_pct') if summary.get('average_return_pct') is not None else '-'}%",
            f"Best recommendation: {summary.get('best_recommendation')}",
            f"Worst recommendation: {summary.get('worst_recommendation')}",
            f"Top missed opportunity: {summary.get('top_missed_opportunity')}",
        ]
    )
    if excel_path:
        lines.append(f"Excel: {excel_path if dry_run else 'attached'}")
    lines.append("")
    lines.append("Recommendation Results:")
    if isinstance(results, pd.DataFrame) and not results.empty:
        for row in results.head(8).to_dict("records"):
            lines.append(
                f"- {row.get('Stock Symbol')}: {row.get('Recommendation Stage')} -> {row.get('Final Result')} | "
                f"close {row.get('Actual Close')} | return {row.get('Actual Return %')}% | {row.get('Reason')}"
            )
    else:
        lines.append("- No recommendation rows available.")
    lines.append("")
    lines.append("Missed Opportunities:")
    if isinstance(missed, pd.DataFrame) and not missed.empty:
        for row in missed.head(5).to_dict("records"):
            lines.append(
                f"- {row.get('Stock Symbol')}: move {row.get('Today Return %')}%, volume {row.get('Volume Change %')}% | "
                f"{row.get('Why Not Selected Code')} | Fix: {row.get('Suggested Fix')}"
            )
    else:
        lines.append("- No top mover missed opportunities found.")
    lines.append("")
    lines.append("Tomorrow Improvement Notes:")
    if isinstance(suggestions, pd.DataFrame) and not suggestions.empty:
        for row in suggestions.head(5).to_dict("records"):
            lines.append(f"- {row.get('Area')}: {row.get('Suggestion')} (auto applied: {row.get('Auto Applied')})")
    if payload.get("learning_telegram_block"):
        lines.extend(["", "Learning Section:", str(payload["learning_telegram_block"])])
    elif payload.get("learning_error"):
        lines.extend(["", f"Learning section unavailable: {payload.get('learning_error')}"])
    lines.extend(["", f"Risk Note: {RISK_NOTE}"])
    return _split_text("\n".join(lines))


def _split_text(text: str, limit: int = 3600) -> list[str]:
    chunks: list[str] = []
    current = ""
    for line in text.splitlines():
        candidate = f"{current}\n{line}".strip() if current else line
        if len(candidate) <= limit:
            current = candidate
            continue
        if current:
            chunks.append(current)
        current = line
    if current:
        chunks.append(current)
    return chunks


def _notification_hash(day: date) -> str:
    return hashlib.sha256(f"end_of_day_prediction_review:{day.isoformat()}".encode("utf-8")).hexdigest()


def send_end_of_day_review_to_telegram(payload: dict[str, Any], *, excel_path: Path | None = None, force: bool = False, db: Session | None = None) -> dict[str, Any]:
    from app.services.telegram_bot import send_private_documents_sync, send_private_message_sync

    day = date.fromisoformat(payload["review_date"])

    def _run(active_db: Session) -> dict[str, Any]:
        existing = active_db.scalar(select(NotificationLog).where(NotificationLog.notification_hash == _notification_hash(day)))
        if existing and not force:
            return {"status": "duplicate_skipped", "sent": False}
        messages = format_end_of_day_telegram(payload, excel_path=excel_path, dry_run=False)
        doc_result = send_private_documents_sync(messages[0], [excel_path] if excel_path and excel_path.exists() and excel_path.stat().st_size < 45 * 1024 * 1024 else [], settings=get_settings())
        sent_chunks = 0
        for chunk in messages[1:]:
            send_private_message_sync(chunk, settings=get_settings())
            sent_chunks += 1
        sent = bool(doc_result.get("sent_messages") or sent_chunks)
        active_db.add(
            NotificationLog(
                notification_hash=_notification_hash(day),
                symbol="ALL",
                notification_type="end_of_day_prediction_review",
                recommendation="REPORT",
                score=_safe_float((payload.get("summary") or {}).get("average_return_pct")),
                source_module="end_of_day_review",
                delivery_status="sent" if sent else "failed",
                cooldown_applied=False,
            )
        )
        active_db.commit()
        return {
            "status": "sent" if sent else "telegram_failed",
            "sent": sent,
            "summary_messages": doc_result.get("sent_messages", 0),
            "sent_documents": doc_result.get("sent_documents", 0),
            "extra_messages": sent_chunks,
            "errors": doc_result.get("errors", []),
        }

    if db is not None:
        return _run(db)
    with sqlite_write_lock():
        with SessionLocal() as active_db:
            return _run(active_db)


def generate_end_of_day_review(
    *,
    target_date: date | None = None,
    persist: bool = True,
    send_telegram: bool = False,
    dry_run: bool = False,
    force_send: bool = False,
    db: Session | None = None,
) -> dict[str, Any]:
    day = target_date or cairo_now().date()

    def _run(active_db: Session) -> dict[str, Any]:
        payload = build_end_of_day_review(target_date=day, persist=persist and not dry_run, db=active_db)
        excel_path = write_end_of_day_excel(payload)
        if payload.get("report_id") and not dry_run:
            report = active_db.get(EndOfDayReviewReport, int(payload["report_id"]))
            if report:
                report.excel_path = str(excel_path)
                active_db.commit()
        telegram = None
        if send_telegram and not dry_run:
            telegram = send_end_of_day_review_to_telegram(payload, excel_path=excel_path, force=force_send, db=active_db)
        return {
            "status": "dry_run" if dry_run else "created",
            "review_date": payload["review_date"],
            "excel_path": str(excel_path),
            "telegram": telegram,
            "telegram_preview": "\n\n".join(format_end_of_day_telegram(payload, excel_path=excel_path, dry_run=True)),
            "summary": payload.get("summary"),
        }

    if db is not None:
        return _run(db)
    init_db(seed=True)
    with sqlite_write_lock():
        with SessionLocal() as active_db:
            return _run(active_db)


def _cli() -> None:
    parser = argparse.ArgumentParser(description="Generate EGX end-of-day prediction and missed-opportunity review.")
    parser.add_argument("--date", default="today", help="today or YYYY-MM-DD")
    parser.add_argument("--send-telegram", action="store_true", help="Send Telegram report and Excel document.")
    parser.add_argument("--force-send", action="store_true", help="Bypass duplicate Telegram protection.")
    parser.add_argument("--dry-run", action="store_true", help="Generate Excel and Telegram preview without sending/persisting.")
    parser.add_argument("--json", action="store_true", help="Print JSON output instead of Telegram preview.")
    args = parser.parse_args()
    result = generate_end_of_day_review(
        target_date=_parse_date(args.date),
        send_telegram=args.send_telegram,
        dry_run=args.dry_run,
        force_send=args.force_send,
    )
    if args.json:
        print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
    else:
        print(result["telegram_preview"])
        print(json.dumps({key: value for key, value in result.items() if key != "telegram_preview"}, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    _cli()
