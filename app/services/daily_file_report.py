"""Daily File Report Generation Module

Generates Excel and PDF reports for EGX recommendations.
Handles morning review, top picks, stocks to avoid, and system analysis.
"""
from __future__ import annotations

import io
import argparse
import json
import logging
import os
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import pandas as pd
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.config import DISCLAIMER, get_settings
from app.models import (
    DailyFileReport,
    FinalStockDecision,
    Opportunity,
    RecommendationItem,
    RecommendationReport,
    Stock,
    StrategyResult,
)

logger = logging.getLogger(__name__)

CAIRO_TZ = ZoneInfo("Africa/Cairo")
REPORTS_DIR = Path(__file__).resolve().parents[2] / "reports"
REPORTS_DIR.mkdir(parents=True, exist_ok=True)


def _fmt(v: Any, d: int = 2) -> str:
    if v is None:
        return "-"
    try:
        return f"{float(v):.{d}f}"
    except (ValueError, TypeError):
        return "-"


def _now_cairo() -> datetime:
    return datetime.now(CAIRO_TZ)


def _today_cairo() -> date:
    return _now_cairo().date()


def generate_excel_report(db: Session, rows: list[dict[str, Any]], filepath: str) -> bool:
    """Generate an Excel report from recommendation rows."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        logger.warning("openpyxl not installed. Installing...")
        os.system("pip install openpyxl -q")
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EGX Daily Report"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    buy_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    watch_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    avoid_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = [
        "Symbol", "Signal", "Score", "Entry Low", "Entry High",
        "Stop Loss", "Target 1", "Target 2", "Target 3",
        "Risk/Reward", "Explanation",
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, item in enumerate(rows, 2):
        signal = (item.get("signal") or "").upper()
        values = [
            item.get("symbol", ""),
            signal,
            _fmt(item.get("final_score")),
            _fmt(item.get("entry_zone_low")),
            _fmt(item.get("entry_zone_high")),
            _fmt(item.get("stop_loss")),
            _fmt(item.get("target_1")),
            _fmt(item.get("target_2")),
            _fmt(item.get("target_3")),
            _fmt(item.get("risk_reward")),
            item.get("explanation", ""),
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if signal in ("BUY", "STRONG BUY"):
                cell.fill = buy_fill
            elif signal in ("WATCH", "WEAK BUY"):
                cell.fill = watch_fill
            elif signal in ("AVOID", "SELL"):
                cell.fill = avoid_fill

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 12
    ws.column_dimensions["K"].width = 60

    wb.save(filepath)
    logger.info("Excel report saved: %s", filepath)
    return True


def write_excel_report(report_data: dict[str, Any], path: str | Path) -> Path:
    """Write the professional multi-sheet daily report workbook.

    This is the stable public helper used by tests and dashboard exports. It
    accepts the data shape returned by collect_report_data().
    """
    from openpyxl.styles import Font, PatternFill

    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    sheet_map = {
        "summary": "Summary",
        "comparison_summary": "Comparison Summary",
        "recommendation_vs_actual": "Recommendation vs Actual",
        "stock_by_stock_comparison": "Stock by Stock Comparison",
        "accuracy_by_stage": "Accuracy by Stage",
        "accuracy_by_strategy": "Accuracy by Strategy",
        "missed_not_evaluated": "Missed - Not Evaluated",
        "best_worst_trades": "Best & Worst Trades",
        "data_quality_check": "Data Quality Check",
        "top_recommendations": "Top 5 Recommendations",
        "audit": "Audit Result",
        "market_evaluation": "Market Evaluation",
        "data_warnings": "Data Warnings",
        "why_failed": "Why Failed",
        "what_next": "What To Improve",
        "repeated_recommendations": "Repeated Recs",
        "errors": "Errors",
        "backtest": "Backtest Summary",
        "telegram": "Telegram Summary",
    }
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for key, sheet_name in sheet_map.items():
            value = report_data.get(key)
            df = value if isinstance(value, pd.DataFrame) else pd.DataFrame(value or [])
            if df.empty:
                df = pd.DataFrame([{"Status": "No data available"}])
            df.to_excel(writer, index=False, sheet_name=sheet_name)

        workbook = writer.book
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for worksheet in workbook.worksheets:
            worksheet.freeze_panes = "A2"
            max_column = worksheet.max_column
            max_row = worksheet.max_row
            if max_column and max_row:
                worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
            header_lookup = {str(cell.value or ""): idx + 1 for idx, cell in enumerate(worksheet[1])}
            result_col = header_lookup.get("result") or header_lookup.get("Result")
            status_col = header_lookup.get("evaluation_status") or header_lookup.get("Evaluation Status")
            quality_label_col = header_lookup.get("final_quality") or header_lookup.get("Final Quality")
            return_col = header_lookup.get("actual_return_pct") or header_lookup.get("Actual Return %")
            quality_col = header_lookup.get("evaluation_quality") or header_lookup.get("Evaluation Quality")
            if result_col or quality_col or status_col or quality_label_col or return_col:
                result_fills = {
                    "GOOD_CALL": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
                    "OPEN_PROFIT": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
                    "BAD_CALL": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
                    "OPEN_LOSS": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
                    "STOP_LOSS_HIT": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
                    "NO_ENTRY": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
                    "DATA_PROBLEM": PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"),
                }
                status_fills = {
                    "TARGET_HIT": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
                    "EVALUATED": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
                    "STOP_HIT": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
                    "EXPIRED": PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid"),
                    "NOT_EVALUATED": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
                    "ENTRY_NOT_REACHED": PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid"),
                    "DATA_MISSING": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
                }
                final_quality_fills = {
                    "EXCELLENT": PatternFill(start_color="00B050", end_color="00B050", fill_type="solid"),
                    "GOOD": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
                    "WEAK": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
                    "BAD": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
                    "NOT EVALUATED": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
                }
                quality_fills = {
                    "HIGH_INTRADAY": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
                    "MEDIUM_DAILY": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
                    "LOW_MISSING_DATA": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
                    "NOT_EVALUATED": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
                }
                for row_idx in range(2, worksheet.max_row + 1):
                    if result_col:
                        result_value = str(worksheet.cell(row=row_idx, column=result_col).value or "").upper()
                        fill = result_fills.get(result_value)
                        if fill:
                            worksheet.cell(row=row_idx, column=result_col).fill = fill
                    if status_col:
                        status_value = str(worksheet.cell(row=row_idx, column=status_col).value or "").upper()
                        fill = status_fills.get(status_value)
                        if fill:
                            worksheet.cell(row=row_idx, column=status_col).fill = fill
                    if quality_label_col:
                        quality_label_value = str(worksheet.cell(row=row_idx, column=quality_label_col).value or "").upper()
                        fill = final_quality_fills.get(quality_label_value)
                        if fill:
                            worksheet.cell(row=row_idx, column=quality_label_col).fill = fill
                    if quality_col:
                        quality_value = str(worksheet.cell(row=row_idx, column=quality_col).value or "").upper()
                        fill = quality_fills.get(quality_value)
                        if fill:
                            worksheet.cell(row=row_idx, column=quality_col).fill = fill
                    if return_col:
                        try:
                            return_value = float(worksheet.cell(row=row_idx, column=return_col).value)
                        except Exception:
                            return_value = None
                        if return_value is not None:
                            worksheet.cell(row=row_idx, column=return_col).fill = (
                                PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                                if return_value >= 0
                                else PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                            )
            for column_cells in worksheet.columns:
                values = [str(cell.value or "") for cell in column_cells]
                width = min(48, max(12, max(len(value) for value in values) + 2))
                worksheet.column_dimensions[column_cells[0].column_letter].width = width
        _add_report_charts(workbook)
    logger.info("Multi-sheet Excel report saved: %s", output)
    return output


def _add_report_charts(workbook: Any) -> None:
    try:
        from openpyxl.chart import BarChart, Reference
    except Exception:
        return

    def header_col(ws: Any, name: str) -> int | None:
        for idx, cell in enumerate(ws[1], start=1):
            if str(cell.value or "").strip() == name:
                return idx
        return None

    def metric_row(ws: Any, label: str) -> int | None:
        for row_idx in range(2, ws.max_row + 1):
            if str(ws.cell(row=row_idx, column=1).value or "").strip() == label:
                return row_idx
        return None

    def add_bar(ws: Any, *, title: str, data_col: int, cat_col: int, min_row: int, max_row: int, anchor: str) -> None:
        if max_row < min_row:
            return
        try:
            chart = BarChart()
            chart.type = "bar"
            chart.style = 10
            chart.title = title
            chart.y_axis.title = ""
            chart.x_axis.title = ""
            data = Reference(ws, min_col=data_col, min_row=min_row, max_row=max_row)
            cats = Reference(ws, min_col=cat_col, min_row=min_row, max_row=max_row)
            chart.add_data(data, titles_from_data=False)
            chart.set_categories(cats)
            chart.height = 7
            chart.width = 12
            ws.add_chart(chart, anchor)
        except Exception:
            logger.debug("Failed to add workbook chart %s", title, exc_info=True)

    if "Accuracy by Stage" in workbook.sheetnames:
        ws = workbook["Accuracy by Stage"]
        stage_col = header_col(ws, "stage")
        win_col = header_col(ws, "win_rate_pct")
        avg_col = header_col(ws, "avg_return_pct")
        if stage_col and win_col and ws.max_row >= 2:
            add_bar(ws, title="Win rate by stage", data_col=win_col, cat_col=stage_col, min_row=2, max_row=ws.max_row, anchor="J2")
        if stage_col and avg_col and ws.max_row >= 2:
            add_bar(ws, title="Average return by stage", data_col=avg_col, cat_col=stage_col, min_row=2, max_row=ws.max_row, anchor="J18")

    if "Comparison Summary" in workbook.sheetnames:
        ws = workbook["Comparison Summary"]
        target_row = metric_row(ws, "Target Hit Count")
        stop_row = metric_row(ws, "Stop Hit Count")
        if target_row and stop_row:
            add_bar(
                ws,
                title="Target hit vs stop hit",
                data_col=2,
                cat_col=1,
                min_row=min(target_row, stop_row),
                max_row=max(target_row, stop_row),
                anchor="E2",
            )

    if "Recommendation vs Actual" in workbook.sheetnames:
        ws = workbook["Recommendation vs Actual"]
        symbol_col = header_col(ws, "symbol")
        return_col = header_col(ws, "actual_return_pct")
        if symbol_col and return_col and ws.max_row >= 2:
            add_bar(
                ws,
                title="Daily recommendation performance",
                data_col=return_col,
                cat_col=symbol_col,
                min_row=2,
                max_row=min(ws.max_row, 16),
                anchor="AA2",
            )

    if "Best & Worst Trades" in workbook.sheetnames:
        ws = workbook["Best & Worst Trades"]
        symbol_col = header_col(ws, "symbol")
        return_col = header_col(ws, "actual_return_pct")
        if symbol_col and return_col and ws.max_row >= 2:
            add_bar(
                ws,
                title="Top 10 best and worst recommendations",
                data_col=return_col,
                cat_col=symbol_col,
                min_row=2,
                max_row=min(ws.max_row, 21),
                anchor="AA2",
            )


def generate_pdf_report(
    db: Session,
    rows: list[dict[str, Any]],
    filepath: str,
    morning_review: dict[str, Any] | None = None,
    report_data: dict[str, Any] | None = None,
) -> bool:
    """Generate a PDF report from recommendation rows."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm, cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    except ImportError:
        logger.warning("reportlab not installed. Installing...")
        os.system("pip install reportlab -q")
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm, cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable

    doc = SimpleDocTemplate(filepath, pagesize=A4)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title2", parent=styles["Title"], fontSize=16, spaceAfter=12)
    heading_style = ParagraphStyle("Heading2Custom", parent=styles["Heading2"], fontSize=13, spaceAfter=8)
    normal_style = ParagraphStyle("NormalCustom", parent=styles["Normal"], fontSize=9, spaceAfter=4)
    small_style = ParagraphStyle("SmallCustom", parent=styles["Normal"], fontSize=8, textColor=colors.grey)

    elements = []
    now_cairo = _now_cairo()
    elements.append(Paragraph(f"EGX Daily Report - {now_cairo.strftime('%Y-%m-%d %H:%M')} Cairo", title_style))
    elements.append(Spacer(1, 6))
    elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#2F5496")))
    elements.append(Spacer(1, 6))

    # Top recommendations
    buy_rows = [r for r in rows if str(r.get("signal") or "").upper() in ("BUY", "STRONG BUY")]
    watch_rows = [r for r in rows if str(r.get("signal") or "").upper() in ("WATCH", "WEAK BUY")]
    avoid_rows = [r for r in rows if str(r.get("signal") or "").upper() in ("AVOID", "SELL", "HIGH_RISK")]

    if buy_rows:
        elements.append(Paragraph("Top BUY Recommendations", heading_style))
        _add_table(elements, buy_rows[:5], normal_style)

    if watch_rows:
        elements.append(Paragraph("Stocks to WATCH", heading_style))
        _add_table(elements, watch_rows[:5], normal_style)

    if avoid_rows:
        elements.append(Paragraph("Stocks to AVOID", heading_style))
        _add_table(elements, avoid_rows[:5], normal_style)

    # Morning review
    if morning_review and morning_review.get("found"):
        elements.append(Spacer(1, 10))
        elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#2F5496")))
        elements.append(Spacer(1, 6))
        elements.append(Paragraph("Morning Recommendation Review", heading_style))

        summary = morning_review.get("summary", {})
        review_text = (
            f"Total: {summary.get('total_recommendations', 0)} | "
            f"Wins: {summary.get('wins', 0)} | Losses: {summary.get('losses', 0)} | "
            f"Win Rate: {summary.get('win_rate_pct', 0)}% | "
            f"Total P&L: {summary.get('total_profit_loss_pct', 0)}%"
        )
        elements.append(Paragraph(review_text, normal_style))
        elements.append(Spacer(1, 4))

        for review in morning_review.get("reviews", []):
            review_line = (
                f"{review['symbol']} ({review['signal']}): "
                f"Entry {review['entry_zone']}, SL {review['stop_loss']}, "
                f"Current {review['current_price']}, P&L {review['profit_loss_pct']}%"
            )
            elements.append(Paragraph(review_line, small_style))

    if report_data:
        comparison_summary = report_data.get("comparison_summary")
        comparison_rows = report_data.get("recommendation_vs_actual")
        if isinstance(comparison_summary, pd.DataFrame) and not comparison_summary.empty:
            elements.append(Spacer(1, 10))
            elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#2F5496")))
            elements.append(Spacer(1, 6))
            elements.append(Paragraph("Recommendation vs Actual Movement", heading_style))
            summary_map = {
                str(row.get("Metric")): row.get("Value")
                for row in comparison_summary.to_dict("records")
            }
            elements.append(
                Paragraph(
                    " | ".join(
                        [
                            f"Compared: {summary_map.get('Recommendations Compared', 0)}",
                            f"Evaluated: {summary_map.get('Evaluated Rows', 0)}",
                            f"Not evaluated: {summary_map.get('Not Evaluated', 0)}",
                            f"Missing: {summary_map.get('Missing Data', 0)}",
                            f"Good: {summary_map.get('Good Calls', '-')}",
                            f"Bad: {summary_map.get('Bad Calls', '-')}",
                            f"Target: {summary_map.get('Target Hit Count', '-')}",
                            f"Stop: {summary_map.get('Stop Hit Count', '-')}",
                            f"Avg Return: {summary_map.get('Average Actual Return %', '-')}",
                        ]
                    ),
                    normal_style,
                )
            )
            diagnosis = summary_map.get("Diagnosis")
            if diagnosis:
                elements.append(Paragraph(f"Diagnosis: {diagnosis}", small_style))
        if isinstance(comparison_rows, pd.DataFrame) and not comparison_rows.empty:
            _add_comparison_table(elements, comparison_rows.head(8).to_dict("records"))

    # System analysis
    elements.append(Spacer(1, 10))
    elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#2F5496")))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph("System Analysis", heading_style))
    elements.append(Paragraph(f"Total stocks tracked: {len(set(r.get('symbol') for r in rows))}", normal_style))
    elements.append(Paragraph(f"BUY signals: {len(buy_rows)} | WATCH: {len(watch_rows)} | AVOID: {len(avoid_rows)}", normal_style))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(f"Disclaimer: {DISCLAIMER}", small_style))

    doc.build(elements)
    logger.info("PDF report saved: %s", filepath)
    return True


def _add_table(elements, rows: list[dict[str, Any]], style):
    from reportlab.platypus import Table, TableStyle, Spacer
    from reportlab.lib import colors

    data = [["Symbol", "Signal", "Score", "Entry", "SL", "Target", "R/R"]]
    for r in rows:
        data.append([
            str(r.get("symbol", "")),
            str(r.get("signal", "")),
            _fmt(r.get("final_score")),
            _fmt(r.get("entry_zone_low")),
            _fmt(r.get("stop_loss")),
            _fmt(r.get("target_1")),
            _fmt(r.get("risk_reward")),
        ])

    table = Table(data, colWidths=[55, 50, 40, 50, 50, 50, 35])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5496")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 6))


def _add_comparison_table(elements, rows: list[dict[str, Any]]) -> None:
    from reportlab.platypus import Table, TableStyle, Spacer
    from reportlab.lib import colors

    data = [["Symbol", "Signal", "Return %", "Status", "Quality", "Stop", "T1", "Root Cause"]]
    for row in rows:
        data.append(
            [
                str(row.get("symbol") or ""),
                str(row.get("recommended_signal") or ""),
                _fmt(row.get("actual_return_pct")),
                str(row.get("evaluation_status") or row.get("result") or ""),
                str(row.get("final_quality") or row.get("evaluation_quality") or ""),
                "Y" if row.get("stop_loss_hit") else "N",
                "Y" if row.get("target_1_hit") else "N",
                str(row.get("root_cause") or "")[:55],
            ]
        )

    table = Table(data, colWidths=[42, 54, 42, 58, 62, 28, 24, 105])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5496")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
            ]
        )
    )
    elements.append(table)
    elements.append(Spacer(1, 6))


write_pdf_report = generate_pdf_report  # backward compat alias


def build_daily_file_report(db: Session | None = None, settings: Any | None = None) -> dict[str, Any]:
    """Build and save Excel and PDF daily reports with morning review."""
    settings = settings or get_settings()
    now_cairo = _now_cairo()
    date_str = now_cairo.strftime("%Y%m%d")
    time_str = now_cairo.strftime("%H%M%S")
    if db is None:
        from app.database import SessionLocal
        db = SessionLocal()
        close_db = True
    else:
        close_db = False

    excel_path = str(REPORTS_DIR / f"EGX_Daily_Report_{date_str}_{time_str}.xlsx")
    pdf_path = str(REPORTS_DIR / f"EGX_Daily_Report_{date_str}_{time_str}.pdf")

    # Get today's opportunities as recommendation items
    opportunities = db.scalars(
        select(Opportunity).order_by(Opportunity.final_score.desc()).limit(50)
    ).all()

    # Fetch morning review
    from app.services.morning_review import review_morning_recommendations
    morning_review = review_morning_recommendations(db, settings)

    rows = []
    for opp in opportunities:
        signal = opp.recommendation
        score = opp.final_score
        entry = opp.entry_price
        target = opp.target_price
        sl = opp.stop_loss

        if signal == "BUY":
            pass
        elif signal == "WATCH":
            pass
        elif signal in ("AVOID", "SELL"):
            pass
        else:
            signal = "NEUTRAL"

        # Calculate risk/reward
        rr = None
        if entry and sl and target and entry != sl:
            potential_loss = abs(entry - sl)
            potential_gain = abs(target - entry)
            if potential_loss > 0:
                rr = round(potential_gain / potential_loss, 2)

        rows.append({
            "symbol": opp.symbol,
            "signal": signal,
            "final_score": score,
            "entry_zone_low": entry,
            "entry_zone_high": entry * 1.005 if entry else None,
            "stop_loss": sl,
            "target_1": target,
            "target_2": target * 1.02 if target else None,
            "target_3": target * 1.05 if target else None,
            "risk_reward": rr,
            "explanation": opp.reason,
        })

    # Generate Excel
    excel_ok = generate_excel_report(db, rows, excel_path) if rows else False
    # Generate PDF
    pdf_ok = generate_pdf_report(db, rows, pdf_path, morning_review) if rows else False

    # Save to database
    report_record = DailyFileReport(
        report_date=now_cairo,
        report_time=now_cairo,
        excel_path=excel_path,
        pdf_path=pdf_path,
        excel_created=excel_ok,
        pdf_created=pdf_ok,
        sent_to_telegram=False,
        status="created",
    )
    db.add(report_record)
    db.commit()
    db.refresh(report_record)

    if close_db:
        db.close()

    return {
        "id": report_record.id,
        "date": now_cairo.isoformat(),
        "excel_path": excel_path if excel_ok else None,
        "pdf_path": pdf_path if pdf_ok else None,
        "excel_created": excel_ok,
        "pdf_created": pdf_ok,
        "items_count": len(rows),
        "morning_review": morning_review,
    }


def send_daily_file_report_to_telegram(db: Session, report_record: DailyFileReport,
                                        settings: Any | None = None) -> bool:
    """Send the daily file report notification to Telegram."""
    settings = settings or get_settings()
    if not settings.telegram_bot_token:
        logger.warning("Telegram not configured; skipping report send.")
        return False

    message = (
        "EGX Daily File Report\n"
        f"Date: {report_record.report_date.strftime('%Y-%m-%d %H:%M')} (Cairo)\n"
        f"Excel: {'Yes' if report_record.excel_created else 'No'}\n"
        f"PDF: {'Yes' if report_record.pdf_created else 'No'}\n\n"
        f"Disclaimer: {DISCLAIMER}"
    )

    try:
        from app.services.telegram_bot import send_private_documents_sync

        paths = []
        if report_record.excel_created and report_record.excel_path and os.path.exists(report_record.excel_path):
            paths.append(report_record.excel_path)
        if report_record.pdf_created and report_record.pdf_path and os.path.exists(report_record.pdf_path):
            paths.append(report_record.pdf_path)
        result = send_private_documents_sync(message, paths, settings=settings)
        report_record.sent_to_telegram = bool(result.get("sent_messages"))
        if not report_record.sent_to_telegram:
            report_record.error_message = "; ".join(str(item) for item in result.get("errors") or []) or "No approved Telegram subscriber received the report."
            report_record.status = "telegram_failed"
        db.commit()
        return bool(report_record.sent_to_telegram)
    except Exception as exc:
        logger.exception("Failed to send Telegram report: %s", exc)
        report_record.error_message = str(exc)
        report_record.status = "telegram_failed"
        db.commit()
        return False


# --- Backward-compatible exports for dashboard/pages/reports_center.py ---

REPORT_DIR = REPORTS_DIR  # dashboard imports REPORT_DIR


def _existing_file_report(db: Session, day_: date) -> DailyFileReport | None:
    """Check if a report already exists for the given date."""
    from sqlalchemy import func as sa_func
    return db.scalar(
        select(DailyFileReport).where(
            sa_func.date(DailyFileReport.report_date) == day_
        ).order_by(DailyFileReport.created_at.desc())
    )


def _safe_number(value: Any) -> float | None:
    if value is None:
        return None
    try:
        return float(value)
    except Exception:
        return None


def _is_evaluated_status(status: Any) -> bool:
    return str(status or "").upper() not in {"NOT_EVALUATED", "DATA_MISSING", "ENTRY_NOT_REACHED", ""}


def _safe_avg(values: list[float]) -> float | None:
    return round(sum(values) / len(values), 2) if values else None


def _comparison_rows(audit: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in audit.get("items") or []:
        details = row.get("details") if isinstance(row.get("details"), dict) else {}
        status = row.get("evaluation_status") or details.get("evaluation_status") or "NOT_EVALUATED"
        strategy_source = row.get("strategy_source") or details.get("strategy_source") or row.get("report_type")
        rows.append(
            {
                "report_type": row.get("report_type"),
                "recommendation_datetime": row.get("report_time"),
                "report_time": row.get("report_time"),
                "symbol": row.get("symbol"),
                "recommended_action_stage": row.get("recommended_signal"),
                "recommended_signal": row.get("recommended_signal"),
                "strategy_source": strategy_source,
                "final_score": row.get("final_score"),
                "entry_zone": row.get("entry_zone"),
                "entry_zone_low": row.get("entry_zone_low"),
                "entry_zone_high": row.get("entry_zone_high"),
                "actual_entry_price": row.get("actual_entry_price"),
                "entry_touched": details.get("entry_touched"),
                "stop_loss": row.get("stop_loss"),
                "stop_loss_hit": details.get("stop_loss_hit"),
                "target_1": row.get("target_1"),
                "target_1_hit": details.get("target_1_hit"),
                "target_2": row.get("target_2"),
                "target_2_hit": details.get("target_2_hit"),
                "target_3": row.get("target_3"),
                "target_3_hit": details.get("target_3_hit"),
                "signal_price": row.get("signal_price") or details.get("signal_price"),
                "next_available_open": row.get("next_available_open") or details.get("next_available_open"),
                "highest_price_after_signal": row.get("highest_price_after_signal") or row.get("max_price_after_signal"),
                "lowest_price_after_signal": row.get("lowest_price_after_signal") or row.get("min_price_after_signal"),
                "max_price_after_signal": row.get("max_price_after_signal"),
                "min_price_after_signal": row.get("min_price_after_signal"),
                "latest_close": row.get("latest_close") or details.get("close_after_recommendation"),
                "close_after_recommendation": details.get("close_after_recommendation"),
                "actual_return_pct": row.get("actual_return"),
                "max_favorable_move_pct": row.get("max_favorable_move_pct") or details.get("max_favorable_move_pct"),
                "max_adverse_move_pct": row.get("max_adverse_move_pct") or details.get("max_adverse_move_pct"),
                "max_drawdown_after_entry_pct": row.get("max_drawdown_after_entry"),
                "max_profit_after_entry_pct": row.get("max_profit_after_entry"),
                "time_to_target_minutes": row.get("time_to_target_minutes"),
                "time_to_stop_minutes": row.get("time_to_stop_minutes"),
                "days_evaluated": row.get("days_evaluated") or details.get("days_evaluated"),
                "evaluation_status": status,
                "final_quality": row.get("final_quality") or details.get("final_quality"),
                "result": row.get("result"),
                "evaluation_quality": row.get("evaluation_quality"),
                "market_regime_at_signal": row.get("market_regime_at_signal"),
                "trade_permission_at_signal": row.get("trade_permission_at_signal"),
                "should_trade_yes_no": row.get("should_trade_yes_no"),
                "mistake_type": row.get("mistake_type"),
                "not_evaluated_reason": row.get("not_evaluated_reason"),
                "root_cause": row.get("root_cause"),
                "fix_required": row.get("fix_required"),
            }
        )
    return rows


def _comparison_summary(audit: dict[str, Any]) -> tuple[list[dict[str, Any]], list[str]]:
    items = audit.get("items") or []
    summary = audit.get("summary") or {}
    status_counts: dict[str, int] = {}
    evaluated_returns: list[float] = []
    evaluated_mfe: list[float] = []
    evaluated_mae: list[float] = []
    evaluated_rows: list[dict[str, Any]] = []
    for row in items:
        status = str(row.get("evaluation_status") or "NOT_EVALUATED")
        status_counts[status] = status_counts.get(status, 0) + 1
        if _is_evaluated_status(status):
            evaluated_rows.append(row)
            number = _safe_number(row.get("actual_return"))
            if number is not None:
                evaluated_returns.append(number)
            mfe = _safe_number(row.get("max_favorable_move_pct"))
            mae = _safe_number(row.get("max_adverse_move_pct"))
            if mfe is not None:
                evaluated_mfe.append(mfe)
            if mae is not None:
                evaluated_mae.append(mae)
    evaluated = len(evaluated_rows)
    avg_return = _safe_avg(evaluated_returns)
    avg_mfe = _safe_avg(evaluated_mfe)
    avg_mae = _safe_avg(evaluated_mae)
    best = max(evaluated_rows, key=lambda row: _safe_number(row.get("actual_return")) if _safe_number(row.get("actual_return")) is not None else -999999, default=None)
    worst = min(evaluated_rows, key=lambda row: _safe_number(row.get("actual_return")) if _safe_number(row.get("actual_return")) is not None else 999999, default=None)
    warnings: list[str] = []
    if not items:
        warnings.append("No recommendation reports were found for this date, so recommendation-vs-actual comparison is empty.")
    elif evaluated == 0:
        warnings.append("No recommendations had future candles after the recommendation timestamp. Do not show accuracy.")
    elif evaluated < 5:
        warnings.append("Accuracy is not reliable yet because evaluated sample size is too small.")
    elif any(str(row.get("evaluation_quality")) == "MEDIUM_DAILY" for row in evaluated_rows) and not any(str(row.get("evaluation_quality")) == "HIGH_INTRADAY" for row in evaluated_rows):
        warnings.append("Comparison uses daily candles only. Intraday order of stop/target events is approximate.")
    reliable_accuracy = evaluated >= 5
    rows = [
        {"Metric": "Comparison Date", "Value": audit.get("audit_date")},
        {"Metric": "Recommendations Compared", "Value": len(items)},
        {"Metric": "Evaluated Rows", "Value": evaluated},
        {"Metric": "Not Evaluated", "Value": status_counts.get("NOT_EVALUATED", 0)},
        {"Metric": "Missing Data", "Value": status_counts.get("DATA_MISSING", 0)},
        {"Metric": "Entry Not Reached", "Value": status_counts.get("ENTRY_NOT_REACHED", 0)},
        {"Metric": "Target Hit Count", "Value": summary.get("target_hit", 0) if evaluated else "Not reliable yet"},
        {"Metric": "Stop Hit Count", "Value": summary.get("stop_loss_hit", 0) if evaluated else "Not reliable yet"},
        {"Metric": "Win Rate %", "Value": summary.get("win_rate_pct") if reliable_accuracy else "Accuracy is not reliable yet because evaluated sample size is too small."},
        {"Metric": "Good Calls", "Value": summary.get("good_calls", 0) if reliable_accuracy else "Not reliable yet"},
        {"Metric": "Bad Calls", "Value": summary.get("bad_calls", 0) if reliable_accuracy else "Not reliable yet"},
        {"Metric": "No Entry / Late", "Value": summary.get("no_entry", 0)},
        {"Metric": "Average Actual Return %", "Value": avg_return if evaluated else "Not reliable yet"},
        {"Metric": "Average Max Favorable Move %", "Value": avg_mfe if evaluated else "Not reliable yet"},
        {"Metric": "Average Max Adverse Move %", "Value": avg_mae if evaluated else "Not reliable yet"},
        {"Metric": "Best Recommendation", "Value": f"{best.get('symbol')} ({best.get('actual_return')}%)" if best else "-"},
        {"Metric": "Worst Recommendation", "Value": f"{worst.get('symbol')} ({worst.get('actual_return')}%)" if worst else "-"},
        {"Metric": "Estimated P&L %", "Value": summary.get("estimated_pnl", 0) if evaluated else "Not reliable yet"},
        {"Metric": "TARGET_HIT Rows", "Value": status_counts.get("TARGET_HIT", 0)},
        {"Metric": "STOP_HIT Rows", "Value": status_counts.get("STOP_HIT", 0)},
        {"Metric": "EVALUATED Rows", "Value": status_counts.get("EVALUATED", 0)},
        {"Metric": "EXPIRED Rows", "Value": status_counts.get("EXPIRED", 0)},
        {"Metric": "NOT_EVALUATED Rows", "Value": status_counts.get("NOT_EVALUATED", 0)},
        {"Metric": "ENTRY_NOT_REACHED Rows", "Value": status_counts.get("ENTRY_NOT_REACHED", 0)},
        {"Metric": "DATA_MISSING Rows", "Value": status_counts.get("DATA_MISSING", 0)},
        {"Metric": "Diagnosis", "Value": audit.get("diagnosis")},
    ]
    return rows, warnings


def _breakdown_rows(rows: list[dict[str, Any]], group_key: str, label_name: str) -> list[dict[str, Any]]:
    groups: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        key = str(row.get(group_key) or "Unknown")
        groups.setdefault(key, []).append(row)
    output: list[dict[str, Any]] = []
    for label, group_rows in sorted(groups.items()):
        evaluated = [row for row in group_rows if _is_evaluated_status(row.get("evaluation_status"))]
        returns = [_safe_number(row.get("actual_return_pct")) for row in evaluated]
        returns = [value for value in returns if value is not None]
        wins = [
            row for row in evaluated
            if row.get("evaluation_status") == "TARGET_HIT"
            or str(row.get("final_quality") or "").upper() in {"EXCELLENT", "GOOD"}
        ]
        target_hit = sum(1 for row in evaluated if row.get("evaluation_status") == "TARGET_HIT" or row.get("target_1_hit") or row.get("target_2_hit") or row.get("target_3_hit"))
        stop_hit = sum(1 for row in evaluated if row.get("evaluation_status") == "STOP_HIT" or row.get("stop_loss_hit"))
        output.append(
            {
                label_name: label,
                "total": len(group_rows),
                "evaluated": len(evaluated),
                "target_hit": target_hit,
                "stop_hit": stop_hit,
                "win_rate_pct": round((len(wins) / len(evaluated)) * 100.0, 2) if len(evaluated) >= 5 else None,
                "accuracy_note": "" if len(evaluated) >= 5 else "Not reliable yet; evaluated sample size is below 5.",
                "avg_return_pct": _safe_avg(returns),
            }
        )
    return output


def _comparison_breakdowns(rows: list[dict[str, Any]]) -> dict[str, pd.DataFrame]:
    evaluated = [row for row in rows if _is_evaluated_status(row.get("evaluation_status"))]
    not_ready = [row for row in rows if not _is_evaluated_status(row.get("evaluation_status"))]
    sorted_eval = sorted(
        evaluated,
        key=lambda row: _safe_number(row.get("actual_return_pct")) if _safe_number(row.get("actual_return_pct")) is not None else -999999,
        reverse=True,
    )
    best_worst = sorted_eval[:10] + list(reversed(sorted_eval[-10:])) if sorted_eval else []
    quality_counts: dict[tuple[str, str, str], int] = {}
    for row in rows:
        key = (
            str(row.get("evaluation_status") or "UNKNOWN"),
            str(row.get("evaluation_quality") or "UNKNOWN"),
            str(row.get("not_evaluated_reason") or row.get("root_cause") or "")[:160],
        )
        quality_counts[key] = quality_counts.get(key, 0) + 1
    data_quality = [
        {
            "evaluation_status": status,
            "evaluation_quality": quality,
            "rows": count,
            "reason": reason,
        }
        for (status, quality, reason), count in sorted(quality_counts.items(), key=lambda item: item[1], reverse=True)
    ]
    return {
        "accuracy_by_stage": pd.DataFrame(_breakdown_rows(rows, "recommended_signal", "stage")),
        "accuracy_by_strategy": pd.DataFrame(_breakdown_rows(rows, "strategy_source", "strategy_source")),
        "missed_not_evaluated": pd.DataFrame(not_ready),
        "best_worst_trades": pd.DataFrame(best_worst),
        "data_quality_check": pd.DataFrame(data_quality),
    }


def collect_report_data(db: Session, day_: date) -> dict[str, Any]:
    """Collect report data for a given date."""
    from app.services.morning_review import review_morning_recommendations
    from app.services.daily_loss_audit import build_daily_loss_audit
    from app.services.market_daily_evaluation import evaluate_daily_market
    from app.services.repeated_recommendation_report import build_repeated_recommendation_report
    from app.config import get_settings
    settings = get_settings()
    opportunities = db.scalars(
        select(Opportunity).order_by(Opportunity.final_score.desc()).limit(50)
    ).all()

    morning_review = review_morning_recommendations(db, settings)

    rows = []
    for opp in opportunities:
        signal = opp.recommendation
        if signal == "BUY":
            pass
        elif signal == "WATCH":
            pass
        elif signal in ("AVOID", "SELL"):
            pass
        else:
            signal = "NEUTRAL"
        rows.append({
            "symbol": opp.symbol,
            "signal": signal,
            "final_score": opp.final_score,
            "entry_zone_low": opp.entry_price,
            "entry_zone_high": opp.entry_price * 1.005 if opp.entry_price else None,
            "stop_loss": opp.stop_loss,
            "target_1": opp.target_price,
            "target_2": opp.target_price * 1.02 if opp.target_price else None,
            "target_3": opp.target_price * 1.05 if opp.target_price else None,
            "risk_reward": None,
            "explanation": opp.reason,
        })

    try:
        audit = build_daily_loss_audit(target_date=day_, persist=False, db=db)
    except Exception as exc:
        audit = {"summary": {}, "items": [], "diagnosis": f"Audit failed: {exc}"}
    comparison_rows = _comparison_rows(audit)
    comparison_summary, comparison_warnings = _comparison_summary(audit)
    comparison_breakdowns = _comparison_breakdowns(comparison_rows)
    try:
        market = evaluate_daily_market(db, target_date=day_, persist=True)
    except Exception as exc:
        market = {"market_score": None, "market_regime": "DATA_INSUFFICIENT", "trade_permission": "DATA_INSUFFICIENT", "warnings": [str(exc)], "explanation": "Market evaluation failed."}
    try:
        repeated = build_repeated_recommendation_report(days=1, end_date=day_, persist=False, db=db)
    except Exception:
        repeated = {"rows": [], "summary": {}}
    warnings = list(market.get("warnings") or [])
    warnings.extend(comparison_warnings)
    try:
        from app.services.recommendation_performance import build_performance_frames

        performance_frames = build_performance_frames(db)
    except Exception as exc:
        performance_frames = {
            "stock_by_stock": pd.DataFrame(),
            "accuracy_by_stage": pd.DataFrame(),
            "accuracy_by_strategy": pd.DataFrame(),
            "accuracy_by_telegram_source": pd.DataFrame(),
            "accuracy_by_market_condition": pd.DataFrame(),
        }
        warnings.append(f"Recommendation performance frames failed: {exc}")
    evaluated = sum(1 for row in audit.get("items") or [] if _is_evaluated_status(row.get("evaluation_status")))
    total_audit = len(audit.get("items") or [])
    if evaluated < 5 and total_audit:
        warnings.append("Accuracy is not reliable yet because evaluated sample size is below 5.")
    if not total_audit:
        warnings.append("No recommendation rows were available for audit on this date.")
    summary_rows = [
        {"Metric": "Report Date", "Value": day_.isoformat()},
        {"Metric": "Market Regime", "Value": market.get("market_regime")},
        {"Metric": "Market Score", "Value": market.get("market_score")},
        {"Metric": "Trade Permission", "Value": market.get("trade_permission")},
        {"Metric": "Recommendations Audited", "Value": total_audit},
        {"Metric": "Evaluated Rows", "Value": evaluated},
        {"Metric": "Not Evaluated Rows", "Value": sum(1 for row in comparison_rows if row.get("evaluation_status") == "NOT_EVALUATED")},
        {"Metric": "Entry Not Reached Rows", "Value": sum(1 for row in comparison_rows if row.get("evaluation_status") == "ENTRY_NOT_REACHED")},
        {"Metric": "Missing Data Rows", "Value": sum(1 for row in comparison_rows if row.get("evaluation_status") == "DATA_MISSING")},
        {"Metric": "Comparison Sheet", "Value": "Recommendation vs Actual"},
        {"Metric": "Accuracy Warning", "Value": "Accuracy is not reliable yet because evaluated sample size is below 5." if evaluated < 5 else "Evaluated with available post-signal data."},
    ]
    why_failed = [
        {"Issue": "Market/Data", "Explanation": warning}
        for warning in warnings
    ] or [{"Issue": "None", "Explanation": "No major report warning detected."}]
    what_next = [
        {"Priority": "P0", "Action": "Keep live trading blocked unless market evaluation says TRADE_ALLOWED and safety flags pass."},
        {"Priority": "P1", "Action": "Improve post-signal OHLCV coverage for higher-confidence audits."},
        {"Priority": "P1", "Action": "Review repeated symbols and dedupe report before sending new alerts."},
    ]
    return {
        "summary": pd.DataFrame(summary_rows),
        "comparison_summary": pd.DataFrame(comparison_summary),
        "recommendation_vs_actual": pd.DataFrame(comparison_rows),
        "stock_by_stock_comparison": performance_frames.get("stock_by_stock", pd.DataFrame()),
        "accuracy_by_stage": comparison_breakdowns["accuracy_by_stage"],
        "accuracy_by_strategy": comparison_breakdowns["accuracy_by_strategy"],
        "missed_not_evaluated": comparison_breakdowns["missed_not_evaluated"],
        "best_worst_trades": comparison_breakdowns["best_worst_trades"],
        "data_quality_check": comparison_breakdowns["data_quality_check"],
        "top_recommendations": pd.DataFrame(rows) if rows else pd.DataFrame(),
        "audit": pd.DataFrame([{key: value for key, value in row.items() if key != "details"} for row in audit.get("items") or []]),
        "market_evaluation": pd.DataFrame([market]),
        "data_warnings": pd.DataFrame([{"warning": warning} for warning in warnings]) if warnings else pd.DataFrame(),
        "why_failed": pd.DataFrame(why_failed),
        "what_next": pd.DataFrame(what_next),
        "errors": pd.DataFrame(),
        "backtest": pd.DataFrame(),
        "telegram": pd.DataFrame(),
        "repeated_recommendations": pd.DataFrame(repeated.get("rows") or []),
        "_morning_review": morning_review,
        "_opportunities": rows,
    }


def _persist_file_recommendation_report(
    db: Session,
    *,
    report_time: datetime,
    rows: list[dict[str, Any]],
) -> RecommendationReport | None:
    if not rows:
        return None
    report = RecommendationReport(
        report_type="file_15pm",
        report_time=report_time.replace(tzinfo=None),
        status="created",
        sent_to_telegram=False,
    )
    db.add(report)
    db.flush()
    for row in rows:
        db.add(
            RecommendationItem(
                report_id=report.id,
                symbol=row.get("symbol"),
                company_name=row.get("company_name") or row.get("symbol"),
                final_score=row.get("final_score"),
                telegram_score=row.get("telegram_score"),
                technical_score=row.get("technical_score"),
                strategy_score=row.get("strategy_score"),
                news_score=row.get("news_score"),
                backtest_score=row.get("backtest_score"),
                risk_liquidity_score=row.get("risk_liquidity_score"),
                signal=row.get("signal"),
                entry_zone_low=row.get("entry_zone_low"),
                entry_zone_high=row.get("entry_zone_high"),
                stop_loss=row.get("stop_loss"),
                target_1=row.get("target_1"),
                target_2=row.get("target_2"),
                target_3=row.get("target_3"),
                risk_reward=row.get("risk_reward"),
                explanation=row.get("explanation"),
                details_json={"source": "daily_file_report", "report_time": report_time.isoformat(timespec="seconds")},
            )
        )
    return report


def generate_daily_file_report(
    report_date: str | date | None = None,
    send_telegram: bool = False,
    excel: bool = True,
    pdf: bool = True,
    force: bool = False,
    dry_run: bool = False,
    reports_dir: str | Path | None = None,
    db: Session | None = None,
) -> dict[str, Any]:
    """Generate daily file report with support for dry-run, override paths, and external DB."""
    from app.config import get_settings
    settings = get_settings()

    if db is None:
        from app.database import SessionLocal
        local_db = SessionLocal()
        close_db = True
    else:
        local_db = db
        close_db = False

    if isinstance(report_date, str):
        day_ = date.fromisoformat(report_date)
    elif isinstance(report_date, date):
        day_ = report_date
    else:
        day_ = _today_cairo()

    if not force:
        existing = _existing_file_report(local_db, day_)
        if existing is not None:
            result: dict[str, Any] = {
                "status": "duplicate_skipped",
                "report_id": existing.id,
                "excel_created": existing.excel_created,
                "pdf_created": existing.pdf_created,
                "sent_to_telegram": existing.sent_to_telegram,
            }
            if send_telegram and not existing.sent_to_telegram and existing.excel_created:
                ok = send_daily_file_report_to_telegram(local_db, existing, settings)
                existing.sent_to_telegram = ok
                local_db.commit()
                result["sent_to_telegram"] = ok
                result["status"] = "sent" if ok else "telegram_failed"
            if close_db:
                local_db.close()
            return result

    data = collect_report_data(local_db, day_)
    opportunities_list = data.get("_opportunities") or []
    result: dict[str, Any] = {
        "data_collected": bool(opportunities_list),
        "items_count": len(opportunities_list),
        "morning_review": data.get("_morning_review"),
    }

    if dry_run:
        result["status"] = "dry_run"
        result["excel_created"] = False
        result["pdf_created"] = False
        result["sent_to_telegram"] = False
        if close_db:
            local_db.close()
        return result

    save_dir = Path(reports_dir) if reports_dir else REPORTS_DIR
    save_dir.mkdir(parents=True, exist_ok=True)

    now_cairo = _now_cairo()
    date_str = now_cairo.strftime("%Y%m%d")
    time_str = now_cairo.strftime("%H%M%S")
    excel_path = str(save_dir / f"EGX_Daily_Report_{date_str}_{time_str}.xlsx")
    pdf_path = str(save_dir / f"EGX_Daily_Report_{date_str}_{time_str}.pdf")

    excel_ok = bool(write_excel_report(data, excel_path)) if excel else False
    try:
        if pdf and opportunities_list:
            try:
                pdf_ok = generate_pdf_report(local_db, opportunities_list, pdf_path, data.get("_morning_review"), report_data=data)
            except TypeError as exc:
                if "report_data" not in str(exc):
                    raise
                pdf_ok = generate_pdf_report(local_db, opportunities_list, pdf_path, data.get("_morning_review"))
        else:
            pdf_ok = False
    except Exception as exc:
        logger.warning("PDF report generation failed: %s", exc)
        pdf_ok = False
        result["error_message"] = f"PDF failed: {exc}"

    report_record = DailyFileReport(
        report_date=now_cairo,
        report_time=now_cairo,
        excel_path=excel_path,
        pdf_path=pdf_path,
        excel_created=excel_ok,
        pdf_created=pdf_ok,
        sent_to_telegram=False,
        status="created",
    )
    local_db.add(report_record)
    recommendation_report = _persist_file_recommendation_report(local_db, report_time=now_cairo, rows=opportunities_list)
    local_db.commit()
    local_db.refresh(report_record)

    if send_telegram and excel_ok:
        ok = send_daily_file_report_to_telegram(local_db, report_record, settings)
        report_record.sent_to_telegram = ok
        local_db.commit()
        result["sent_to_telegram"] = ok
        result["status"] = "sent" if ok else "telegram_failed"
    else:
        result["status"] = "success"
        result["sent_to_telegram"] = False

    result["report_id"] = report_record.id
    result["recommendation_report_id"] = recommendation_report.id if recommendation_report is not None else None
    result["excel_created"] = excel_ok
    result["pdf_created"] = pdf_ok
    result["excel_path"] = excel_path if excel_ok else None
    result["pdf_path"] = pdf_path if pdf_ok else None

    if close_db:
        local_db.close()
    return result


def latest_file_reports(db: Session, limit: int = 100) -> list[DailyFileReport]:
    """Backward-compatible wrapper."""
    return list(db.scalars(
        select(DailyFileReport).order_by(DailyFileReport.created_at.desc()).limit(limit)
    ).all())


def send_report_to_telegram(row: DailyFileReport) -> dict[str, Any]:
    """Backward-compatible wrapper."""
    from app.config import get_settings
    from app.database import SessionLocal
    settings = get_settings()
    with SessionLocal() as db2:
        active_row = db2.get(DailyFileReport, row.id)
        if active_row is None:
            return {"sent": False, "sent_messages": 0, "sent_documents": 0, "error": "Report row not found."}
        ok = send_daily_file_report_to_telegram(db2, active_row, settings)
        docs = int(bool(active_row.excel_created and active_row.excel_path)) + int(bool(active_row.pdf_created and active_row.pdf_path))
    return {"sent": ok, "sent_messages": 1 if ok else 0, "sent_documents": docs if ok else 0}


def target_day(iso_date: str) -> str:
    """Backward-compatible wrapper; just returns the ISO date string."""
    return iso_date


def _parse_report_date(value: str | None) -> date | None:
    if not value or value.lower() == "today":
        return _today_cairo()
    return date.fromisoformat(value)


def _cli() -> None:
    parser = argparse.ArgumentParser(description="Generate EGX daily Excel/PDF file report.")
    parser.add_argument("--date", default="today", help="today or YYYY-MM-DD")
    parser.add_argument("--send-telegram", action="store_true", help="Send report notification/documents to Telegram.")
    parser.add_argument("--excel-only", action="store_true", help="Generate Excel only.")
    parser.add_argument("--pdf-only", action="store_true", help="Generate PDF only.")
    parser.add_argument("--dry-run", action="store_true", help="Collect report data without writing files or sending Telegram.")
    parser.add_argument("--force", action="store_true", help="Create a new report even if today's report already exists.")
    args = parser.parse_args()

    excel = not args.pdf_only
    pdf = not args.excel_only
    result = generate_daily_file_report(
        report_date=_parse_report_date(args.date),
        send_telegram=args.send_telegram,
        excel=excel,
        pdf=pdf,
        force=args.force,
        dry_run=args.dry_run,
    )
    print(json.dumps(result, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    _cli()
