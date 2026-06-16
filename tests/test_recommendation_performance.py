from __future__ import annotations

from datetime import date, datetime, timedelta
from pathlib import Path
from types import SimpleNamespace

import pandas as pd
from openpyxl import load_workbook

from app.services import recommendation_performance as perf
from app.services.daily_loss_audit import (
    EVAL_DATA_MISSING,
    EVAL_EVALUATED,
    EVAL_ENTRY_NOT_REACHED,
    EVAL_NOT_EVALUATED,
    EVAL_STOP_HIT,
    EVAL_TARGET_HIT,
)


class FakeDB:
    def __init__(self, duplicate=None) -> None:  # noqa: ANN001
        self.duplicate = duplicate
        self.committed = False

    def scalar(self, *args, **kwargs):  # noqa: ANN002, ANN003
        return self.duplicate

    def commit(self) -> None:
        self.committed = True


def _item() -> SimpleNamespace:
    return SimpleNamespace(
        id=1,
        symbol="COMI",
        company_name="Commercial International Bank",
        signal="CONDITIONAL BUY",
        entry_zone_low=10.0,
        entry_zone_high=11.0,
        stop_loss=9.0,
        target_1=12.0,
        target_2=13.0,
        target_3=14.0,
        details_json={"source": "daily_report", "telegram_source": "channel_a"},
    )


def _report() -> SimpleNamespace:
    return SimpleNamespace(id=10, report_type="evening", report_time=datetime(2026, 6, 7, 21, 0, 0))


def _candles(*rows: tuple[str, float, float, float, float]) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {"datetime": value[0], "open": value[1], "high": value[2], "low": value[3], "close": value[4], "volume": 1000}
            for value in rows
        ]
    )


def _sample_frames() -> dict[str, pd.DataFrame]:
    stock = pd.DataFrame(
        [
            {
                "Report Date": "2026-06-07",
                "Recommendation Date": "2026-06-07 21:00:00",
                "Stock Symbol": "COMI",
                "Stock Name": "Commercial International Bank",
                "Recommendation Stage": "CONDITIONAL BUY",
                "Strategy": "daily_report",
                "Telegram Source": "channel_a",
                "Market Condition": "NEUTRAL",
                "Entry From": 10.0,
                "Entry To": 11.0,
                "Stop Loss": 9.0,
                "Target 1": 12.0,
                "Target 2": 13.0,
                "Signal Price": 10.5,
                "Latest Close": 12.2,
                "Highest After Signal": 12.4,
                "Lowest After Signal": 10.2,
                "Actual Return %": 10.9,
                "Max Favorable Move %": 12.7,
                "Max Adverse Move %": -7.3,
                "Days Evaluated": 1,
                "Status": EVAL_TARGET_HIT,
                "Quality": "Excellent",
                "Notes": "Target reached before invalidation.",
                "Evaluated At": datetime(2026, 6, 8, 15, 30, 0),
            },
            {
                "Report Date": "2026-06-08",
                "Recommendation Date": "2026-06-08 21:00:00",
                "Stock Symbol": "AALR",
                "Stock Name": "AALR",
                "Recommendation Stage": "STRONG BUY",
                "Strategy": "daily_report",
                "Telegram Source": "channel_b",
                "Market Condition": "BULLISH",
                "Entry From": 1.2,
                "Entry To": 1.25,
                "Stop Loss": 1.15,
                "Target 1": 1.35,
                "Target 2": 1.42,
                "Signal Price": 1.23,
                "Latest Close": None,
                "Highest After Signal": None,
                "Lowest After Signal": None,
                "Actual Return %": None,
                "Max Favorable Move %": None,
                "Max Adverse Move %": None,
                "Days Evaluated": 0,
                "Status": EVAL_NOT_EVALUATED,
                "Quality": "Not Evaluated",
                "Notes": "No future candle exists after the recommendation timestamp yet.",
                "Evaluated At": datetime(2026, 6, 8, 15, 30, 0),
            },
        ]
    )
    summary = perf.summarize_evaluations(perf._summary_records_from_frame(stock))
    return {
        "summary": pd.DataFrame([{"Metric": key, "Value": value} for key, value in summary.items()]),
        "stock_by_stock": stock,
        "accuracy_by_stage": perf._group_accuracy(stock, "Recommendation Stage"),
        "accuracy_by_strategy": perf._group_accuracy(stock, "Strategy"),
        "accuracy_by_telegram_source": perf._group_accuracy(stock, "Telegram Source"),
        "accuracy_by_market_condition": perf._group_accuracy(stock, "Market Condition"),
    }


def test_old_recommendation_becomes_target_hit_when_future_candle_appears(monkeypatch) -> None:
    future = _candles(("2026-06-08 10:00:00", 10.6, 12.2, 10.4, 12.0))
    monkeypatch.setattr(perf, "_candles_after_recommendation", lambda *args, **kwargs: (future, "MEDIUM_DAILY", "1d"))
    monkeypatch.setattr(perf, "_latest_signal_price_before", lambda *args, **kwargs: 10.5)
    audit, horizons, _future, quality, _timeframe = perf._evaluate_path(FakeDB(), _item(), _report())
    assert audit.evaluation_status == EVAL_TARGET_HIT
    assert audit.target_1_hit is True
    assert quality == "MEDIUM_DAILY"
    assert horizons["1d"]["status"] == EVAL_TARGET_HIT


def test_stop_hit_updates_correctly(monkeypatch) -> None:
    future = _candles(("2026-06-08 10:00:00", 10.7, 10.9, 8.8, 9.1))
    monkeypatch.setattr(perf, "_candles_after_recommendation", lambda *args, **kwargs: (future, "MEDIUM_DAILY", "1d"))
    monkeypatch.setattr(perf, "_latest_signal_price_before", lambda *args, **kwargs: 10.5)
    audit, _horizons, _future, _quality, _timeframe = perf._evaluate_path(FakeDB(), _item(), _report())
    assert audit.evaluation_status == EVAL_STOP_HIT
    assert audit.stop_loss_hit is True


def test_open_recommendation_remains_evaluated_if_target_and_stop_not_hit(monkeypatch) -> None:
    future = _candles(("2026-06-08 10:00:00", 10.7, 11.5, 10.2, 11.2))
    monkeypatch.setattr(perf, "_candles_after_recommendation", lambda *args, **kwargs: (future, "MEDIUM_DAILY", "1d"))
    monkeypatch.setattr(perf, "_latest_signal_price_before", lambda *args, **kwargs: 10.5)
    audit, _horizons, _future, _quality, _timeframe = perf._evaluate_path(FakeDB(), _item(), _report())
    assert audit.evaluation_status == EVAL_EVALUATED
    assert audit.target_1_hit is False
    assert audit.stop_loss_hit is False


def test_accuracy_excludes_missing_and_not_evaluated_rows() -> None:
    rows = [
        {"symbol": "A", "final_status": EVAL_TARGET_HIT, "actual_return_pct": 4.0},
        {"symbol": "B", "final_status": EVAL_STOP_HIT, "actual_return_pct": -2.0},
        {"symbol": "C", "final_status": EVAL_NOT_EVALUATED, "actual_return_pct": 99.0},
        {"symbol": "D", "final_status": EVAL_DATA_MISSING, "actual_return_pct": -99.0},
        {"symbol": "E", "final_status": EVAL_ENTRY_NOT_REACHED, "actual_return_pct": 8.0},
    ]
    summary = perf.summarize_evaluations(rows)
    assert summary["evaluated_recommendations"] == 2
    assert summary["not_evaluated_recommendations"] == 1
    assert summary["missing_data_recommendations"] == 1
    assert summary["entry_not_reached_recommendations"] == 1
    assert summary["win_rate_pct"] is None
    assert "below 5" in summary["accuracy_note"]


def test_stock_by_stock_messages_include_every_recommended_stock() -> None:
    messages = perf.stock_by_stock_messages(_sample_frames()["stock_by_stock"])
    combined = "\n".join(messages)
    assert "COMI" in combined
    assert "AALR" in combined
    assert "Not evaluated yet" in combined or "No future candle" in combined


def test_long_telegram_messages_are_split_safely() -> None:
    chunks = perf.split_telegram_messages(["A" * 100, "B" * 4000, "C" * 100], limit=500)
    assert chunks
    assert all(len(chunk) <= 500 for chunk in chunks)


def test_performance_excel_contains_stock_by_stock_sheet(tmp_path: Path) -> None:
    path = tmp_path / "performance.xlsx"
    perf.write_performance_excel(_sample_frames(), path)
    workbook = load_workbook(path)
    assert "Stock by Stock Comparison" in workbook.sheetnames
    headers = [cell.value for cell in workbook["Stock by Stock Comparison"][1]]
    assert "Stock Symbol" in headers
    assert "Status" in headers


def test_duplicate_telegram_report_is_blocked() -> None:
    result = perf.send_performance_report_to_telegram(as_of_date=date(2026, 6, 8), db=FakeDB(duplicate=object()))
    assert result["status"] == "duplicate_skipped"
    assert result["sent"] is False


def test_telegram_sends_summary_document_and_stock_chunks(monkeypatch, tmp_path: Path) -> None:
    sent = {"documents": 0, "messages": []}
    xlsx = tmp_path / "performance.xlsx"
    xlsx.write_text("placeholder", encoding="utf-8")

    import app.services.telegram_bot as telegram_bot

    monkeypatch.setattr(perf, "build_performance_frames", lambda db: _sample_frames())
    monkeypatch.setattr(perf, "generate_performance_excel", lambda **kwargs: xlsx)
    monkeypatch.setattr(perf, "_duplicate_sent", lambda db, as_of_date: False)
    monkeypatch.setattr(perf, "_mark_sent", lambda *args, **kwargs: None)

    def fake_docs(text, document_paths, settings=None):  # noqa: ANN001
        sent["documents"] += len(document_paths)
        sent["messages"].append(text)
        return {"sent_messages": 1, "sent_documents": len(document_paths), "errors": []}

    def fake_message(text, settings=None):  # noqa: ANN001
        sent["messages"].append(text)
        return {"sent_messages": 1, "errors": []}

    monkeypatch.setattr(telegram_bot, "send_private_documents_sync", fake_docs)
    monkeypatch.setattr(telegram_bot, "send_private_message_sync", fake_message)

    result = perf.send_performance_report_to_telegram(as_of_date=date(2026, 6, 8), db=FakeDB())
    assert result["sent"] is True
    assert sent["documents"] == 1
    combined = "\n".join(sent["messages"])
    assert "EGX Recommendation Performance Report" in combined
    assert "COMI" in combined
    assert "AALR" in combined
    assert "Win rate: Accuracy is not reliable" in combined
