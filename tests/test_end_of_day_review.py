from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from app.services import end_of_day_review as eod
from app.services import daily_loss_audit
from app.services.daily_loss_audit import EVAL_ENTRY_NOT_REACHED, EVAL_NOT_EVALUATED


def _payload() -> dict:
    results = pd.DataFrame(
        [
            {
                "Stock Symbol": "COMI",
                "Recommendation Stage": "BUY",
                "Final Result": "TARGET_HIT",
                "Target Hit": True,
                "Stop Hit": False,
                "Actual Return %": 4.2,
                "Quality Rating": "Excellent",
                "Reason": "Target reached.",
                "Technical Score": 82,
                "Financial Score": 65,
                "News Score": 70,
                "Telegram Score": 76,
                "Strategy Score": 75,
                "Liquidity Score": 88,
            },
            {
                "Stock Symbol": "AALR",
                "Recommendation Stage": "WATCH",
                "Final Result": EVAL_ENTRY_NOT_REACHED,
                "Target Hit": False,
                "Stop Hit": False,
                "Actual Return %": 3.1,
                "Quality Rating": "Not Evaluated",
                "Reason": "Entry was not reached.",
                "Technical Score": 70,
                "Financial Score": 50,
                "News Score": 50,
                "Telegram Score": 60,
                "Strategy Score": 55,
                "Liquidity Score": 70,
            },
        ]
    )
    missed = pd.DataFrame(
        [
            {
                "Stock Symbol": "XXXX",
                "Stock Name": "Missed Stock",
                "Today Return %": 8.4,
                "Volume Change %": 120,
                "Value Traded": 2_000_000,
                "Why Not Selected Code": "LATE_BREAKOUT",
                "Why Not Selected": "LATE_BREAKOUT: breakout happened after scan.",
                "Suggested Fix": "Add intraday re-scans.",
                "Technical Score": 78,
                "Telegram Score": 40,
                "News Score": 50,
                "Financial Score": 50,
                "Liquidity Score": 80,
                "Final Score": 68,
                "Selected Today": False,
            }
        ]
    )
    summary = eod._summary_from_results(results.to_dict("records"), missed.to_dict("records"))
    suggestions = eod._strategy_suggestions(results.to_dict("records"), missed.to_dict("records"), missed.to_dict("records"))
    return {
        "review_date": "2026-06-14",
        "summary": summary,
        "market_evaluation": {"market_score": 65, "trade_permission": "WATCH_ONLY"},
        "daily_prediction_review": pd.DataFrame([summary]),
        "recommendation_results": results,
        "missed_opportunities": missed,
        "why_not_selected": missed,
        "top_movers_analysis": missed,
        "strategy_improvement_suggestions": pd.DataFrame(suggestions),
        "score_breakdown": missed,
        "data_quality_issues": pd.DataFrame([{"Stock Symbol": "COMI", "Status": "OK"}]),
        "telegram_vs_actual": pd.DataFrame([{"Stock Symbol": "COMI", "Telegram Score": 76, "Actual Return %": 4.2, "Selected": True}]),
        "technical_vs_actual": pd.DataFrame([{"Stock Symbol": "COMI", "Technical Score": 82, "Actual Return %": 4.2, "Selected": True}]),
        "financial_vs_actual": pd.DataFrame([{"Stock Symbol": "COMI", "Financial Score": 65, "Actual Return %": 4.2, "Selected": True}]),
        "news_vs_actual": pd.DataFrame([{"Stock Symbol": "COMI", "News Score": 70, "Actual Return %": 4.2, "Selected": True}]),
        "tomorrow_watchlist": missed,
        "risk_note": "System-generated analysis, not financial advice.",
    }


def test_recommended_stock_summary_excludes_entry_not_reached_from_accuracy() -> None:
    summary = _payload()["summary"]
    assert summary["total_recommendations"] == 2
    assert summary["evaluated_recommendations"] == 1
    assert summary["entry_not_reached"] == 1
    assert summary["win_rate_pct"] is None


def test_daily_move_from_ohlcv_calculates_top_mover_fields() -> None:
    df = pd.DataFrame(
        [
            {"datetime": "2026-06-13", "open": 10, "high": 10.4, "low": 9.8, "close": 10, "volume": 1000},
            {"datetime": "2026-06-14", "open": 10.2, "high": 11.2, "low": 10.1, "close": 11, "volume": 3000},
        ]
    )
    move = eod.daily_move_from_ohlcv(df, date(2026, 6, 14))
    assert move["status"] == "OK"
    assert move["return_pct"] == 10.0
    assert move["volume_change_pct"] == 200.0
    assert move["value_traded"] == 33000


def test_why_not_selected_low_liquidity_is_flagged() -> None:
    row = {"technical_score": 80, "strategy_score": 80, "news_score": 60, "telegram_score": 60, "risk_liquidity_score": 30, "final_score": 75, "risk_reward": 2}
    move = {"status": "OK", "value_traded": 100_000, "return_pct": 6, "volume_change_pct": 90}
    assert eod.classify_why_not_selected(row, {"trade_permission": "TRADE_ALLOWED"}, move) == "LOW_LIQUIDITY"


def test_why_not_selected_late_breakout_is_detected() -> None:
    row = {"technical_score": 72, "strategy_score": 62, "news_score": 50, "telegram_score": 50, "risk_liquidity_score": 75, "final_score": 64, "risk_reward": 2}
    move = {"status": "OK", "value_traded": 2_000_000, "return_pct": 7.2, "volume_change_pct": 110}
    assert eod.classify_why_not_selected(row, {"trade_permission": "TRADE_ALLOWED"}, move) == "LATE_BREAKOUT"


def test_strategy_suggestions_are_not_auto_applied() -> None:
    suggestions = eod._strategy_suggestions([], [{"Why Not Selected Code": "LATE_BREAKOUT"}], [])
    assert suggestions
    assert all(item["Auto Applied"] == "No" for item in suggestions)


def test_telegram_end_of_day_summary_contains_missed_and_entry_not_reached() -> None:
    text = "\n".join(eod.format_end_of_day_telegram(_payload(), dry_run=True))
    assert "Entry not reached: 1" in text
    assert "Missed Opportunities" in text
    assert "XXXX" in text
    assert "Accuracy is not reliable" in text


def test_excel_contains_missed_opportunities_sheet(tmp_path: Path) -> None:
    path = eod.write_end_of_day_excel(_payload(), tmp_path / "eod.xlsx")
    workbook = load_workbook(path)
    assert "Missed Opportunities" in workbook.sheetnames
    assert "Why Not Selected" in workbook.sheetnames
    assert "Strategy Improvement Suggestion" in workbook.sheetnames
    assert "Tomorrow Watchlist" in workbook.sheetnames


def test_same_day_daily_candle_is_not_used_as_future_evaluation(monkeypatch) -> None:
    same_day = pd.DataFrame(
        [{"datetime": "2026-06-14", "open": 10, "high": 12, "low": 9, "close": 11, "volume": 1000}]
    )

    def fake_get_ohlcv(db, symbol, timeframe=None, limit=300):  # noqa: ANN001
        if timeframe in {"15m", "30m", "1h", "4h"}:
            return pd.DataFrame()
        return same_day

    monkeypatch.setattr(daily_loss_audit, "get_ohlcv", fake_get_ohlcv)
    candles, quality, timeframe = daily_loss_audit._candles_after_recommendation(
        object(),
        "COMI",
        datetime(2026, 6, 14, 9, 0, 0),
        date(2026, 6, 14),
    )
    assert candles.empty
    assert quality == "NOT_EVALUATED"
    assert timeframe == "none"


def test_entry_not_reached_status_is_non_accuracy() -> None:
    summary = eod._summary_from_results(
        [{"Stock Symbol": "A", "Final Result": EVAL_ENTRY_NOT_REACHED, "Actual Return %": 5.0}],
        [],
    )
    assert summary["evaluated_recommendations"] == 0
    assert summary["win_rate_pct"] is None
