from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from app.services.daily_file_report import write_excel_report
from tests.test_daily_file_report import sample_report_data


def test_excel_report_is_created_with_required_sheets(tmp_path: Path) -> None:
    path = tmp_path / "EGX_Daily_Report_2026-06-07.xlsx"
    write_excel_report(sample_report_data(), path)
    assert path.exists()
    workbook = load_workbook(path)
    assert "Summary" in workbook.sheetnames
    assert "Comparison Summary" in workbook.sheetnames
    assert "Recommendation vs Actual" in workbook.sheetnames
    assert "Stock by Stock Comparison" in workbook.sheetnames
    assert "Accuracy by Stage" in workbook.sheetnames
    assert "Accuracy by Strategy" in workbook.sheetnames
    assert "Missed - Not Evaluated" in workbook.sheetnames
    assert "Best & Worst Trades" in workbook.sheetnames
    assert "Data Quality Check" in workbook.sheetnames
    assert "Top 5 Recommendations" in workbook.sheetnames
    assert "Audit Result" in workbook.sheetnames
    assert "Errors" in workbook.sheetnames
    assert workbook["Top 5 Recommendations"].freeze_panes == "A2"
    assert workbook["Recommendation vs Actual"].freeze_panes == "A2"
